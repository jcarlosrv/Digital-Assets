import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import networkx as nx
from pyvis.network import Network
import streamlit.components.v1 as components
import tempfile

# ── Page config ──
st.set_page_config(
    page_title="Chicago Digital Asset Model",
    page_icon="🏙️",
    layout="wide",
)

# ── Custom CSS ──
st.markdown("""
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
    * { font-family: 'Inter', sans-serif !important; }
    .block-container { padding-top: 1.5rem; background-color: #F8FAFC; }
    div[data-testid="stMetric"] {
        background: linear-gradient(135deg, #1B3A5F 0%, #0F2744 100%);
        padding: 14px 18px;
        border-radius: 10px;
        color: white;
        border-left: 4px solid #D4A84B;
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
    }
    div[data-testid="stMetric"] label { color: #94A3B8 !important; font-size: 0.8rem !important; }
    div[data-testid="stMetric"] div[data-testid="stMetricValue"] { color: #D4A84B !important; font-size: 1.3rem !important; font-weight: 600 !important; }
    div[data-testid="stMetric"] div[data-testid="stMetricDelta"] { color: #059669 !important; font-size: 0.75rem !important; }
    h1 { color: #0F2744 !important; font-weight: 700 !important; }
    h2, h3 { color: #1B3A5F !important; font-weight: 600 !important; }
    div[data-testid="stTabs"] button { color: #1E293B !important; font-weight: 500 !important; }
    div[data-testid="stTabs"] button[aria-selected="true"] { border-bottom-color: #1A7A72 !important; color: #1A7A72 !important; font-weight: 600 !important; }
    .stDataFrame { border-radius: 8px; box-shadow: 0 1px 4px rgba(0,0,0,0.08); }
    div[data-testid="stExpander"] { border-radius: 8px; border: 1px solid #E2E8F0; }
    /* Sidebar dark navy */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0F2744 0%, #1B3A5F 100%) !important;
    }
    section[data-testid="stSidebar"] * {
        color: #E2E8F0 !important;
    }
    section[data-testid="stSidebar"] div[data-testid="stMetric"] label,
    section[data-testid="stSidebar"] div[data-testid="stMetric"] div[data-testid="stMetricValue"] {
        color: #D4A84B !important;
    }
    section[data-testid="stSidebar"] .stSelectbox label,
    section[data-testid="stSidebar"] .stSlider label,
    section[data-testid="stSidebar"] .stMultiSelect label {
        color: #94A3B8 !important;
        font-weight: 500 !important;
    }
    section[data-testid="stSidebar"] hr {
        border-color: rgba(255,255,255,0.15) !important;
    }
    section[data-testid="stSidebar"] button {
        background-color: #1A7A72 !important;
        color: white !important;
        border: none !important;
        border-radius: 6px !important;
    }
</style>
""", unsafe_allow_html=True)

# ── Format helper ──
def fmt(value):
    """Format large numbers: $1.2B, $45.3M, $120.5K"""
    if abs(value) >= 1e9:
        return f"${value / 1e9:,.1f}B"
    elif abs(value) >= 1e6:
        return f"${value / 1e6:,.1f}M"
    elif abs(value) >= 1e3:
        return f"${value / 1e3:,.1f}K"
    else:
        return f"${value:,.0f}"

# ── Load data ──
@st.cache_data
def load_data():
    import os
    import openpyxl
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(script_dir, "IT_Budget_Chicago_Clean.xlsx")
    # Use openpyxl with data_only to resolve VLOOKUP formulas
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb["Chicago IT budget_clean"]
    data = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        data.append(row)
    columns = ["Sector", "Short Name", "Department", "IT Type", 2021, 2022, 2023, 2024, 2025, 2026]
    df = pd.DataFrame(data, columns=columns)
    years = [2021, 2022, 2023, 2024, 2025, 2026]
    df["IT Type"] = df["IT Type"].str.strip()
    df["Sector"] = df["Sector"].fillna("Other")
    df["Short Name"] = df["Short Name"].fillna(df["Department"])
    # Build short name lookup
    short_name_map = df.drop_duplicates("Department").set_index("Department")["Short Name"].to_dict()
    df_long = df.melt(
        id_vars=["Sector", "Short Name", "Department", "IT Type"],
        value_vars=years,
        var_name="Year",
        value_name="Budget"
    )
    df_long["Budget"] = pd.to_numeric(df_long["Budget"], errors="coerce").fillna(0)
    return df, df_long, short_name_map

df_wide, df, short_name_map = load_data()
all_hist_years = [2021, 2022, 2023, 2024, 2025, 2026]
future_years = [2027, 2028, 2029, 2030, 2031]

# ── Projection helper ──
def project_series(hist_years, hist_values, future_years, multiplier=1.0, noise_pct=0.0, seed=42):
    """Project from last actual value using linear trend slope, adjusted by multiplier. Optional noise."""
    x = np.array(hist_years, dtype=float)
    y = np.array(hist_values, dtype=float)
    nonzero = y > 0
    if nonzero.sum() < 2:
        return [0.0] * len(future_years)
    slope = np.polyfit(x[nonzero], y[nonzero], 1)[0]
    adjusted_slope = slope * multiplier
    last_value = y[-1]
    last_year = x[-1]
    projected = [max(0, last_value + adjusted_slope * (fy - last_year)) for fy in future_years]
    if noise_pct > 0:
        rng = np.random.default_rng(seed)
        for i in range(len(projected)):
            noise = rng.uniform(-noise_pct, noise_pct) * projected[i]
            projected[i] = max(0, projected[i] + noise)
    return projected


def build_pyvis_network(df_overlap, type_col, type_colors, title):
    """Build a pyvis network HTML string for overlap/fragmentation data."""
    sector_colors = {
        "Public Safety": "#8B2942", "Finance and Administration": "#1B3A5F",
        "Infrastructure Services": "#1A7A72", "Community Services": "#D4A84B",
        "City Development": "#D97706", "Legislative and Election": "#0F2744",
        "Regulatory": "#3B82F6",
    }

    G = nx.Graph()

    for _, row in df_overlap.iterrows():
        group = row["Overlap Group"]
        typ = row[type_col]
        overlap_pct = row["Overlap %"] if "Overlap %" in row else row.get("Estimated Overlap (%)", 50)
        sectors = [s.strip() for s in str(row["Sectors Involved"]).split(",")]
        savings = row.get("Potential Savings", 0)

        G.add_node(group, label=group, title=f"<b>{group}</b><br>Type: {typ}<br>Overlap: {overlap_pct:.0f}%<br>Savings: ${savings:,.0f}",
                    color=type_colors.get(typ, "#888"), size=10 + overlap_pct * 0.1, shape="dot", font={"size": 8, "color": "#555"})

        for sector in sectors:
            if not G.has_node(sector):
                G.add_node(sector, label=sector.replace(" and ", " & "), title=f"<b>{sector}</b>",
                           color=sector_colors.get(sector, "#ccc"), size=50, shape="dot",
                           borderWidth=3, font={"size": 14, "color": "#000", "bold": True})
            G.add_edge(sector, group, width=max(1, overlap_pct / 20),
                       color={"color": type_colors.get(typ, "#888"), "opacity": 0.6},
                       title=f"{sector} ↔ {group}<br>Overlap: {overlap_pct:.0f}%")

        # Sector-to-sector implicit edges
        for i in range(len(sectors)):
            for j in range(i + 1, len(sectors)):
                if G.has_edge(sectors[i], sectors[j]):
                    G[sectors[i]][sectors[j]]["width"] += overlap_pct / 30
                else:
                    G.add_edge(sectors[i], sectors[j], width=overlap_pct / 30,
                               color={"color": "#ccc", "opacity": 0.3},
                               title=f"{sectors[i]} ↔ {sectors[j]}", dashes=True)

    net = Network(height="550px", width="100%", bgcolor="#ffffff", font_color="#333")
    net.from_nx(G)
    net.set_options("""
    {"physics": {"forceAtlas2Based": {"gravitationalConstant": -80, "centralGravity": 0.01,
     "springLength": 150, "springConstant": 0.02, "damping": 0.4}, "solver": "forceAtlas2Based",
     "stabilization": {"iterations": 200}},
     "interaction": {"hover": true, "tooltipDelay": 100, "zoomView": true},
     "edges": {"smooth": {"type": "continuous"}}}
    """)

    # Build legend HTML
    type_legend = " ".join(f'<span style="display:inline-block;width:10px;height:10px;background:{c};border-radius:50%;margin-right:3px;"></span>{n}&nbsp;' for n, c in type_colors.items())
    sect_legend = " ".join(f'<span style="display:inline-block;width:10px;height:10px;background:{c};border-radius:50%;margin-right:3px;"></span>{n.replace(" and ", " & ")}&nbsp;' for n, c in sector_colors.items())
    legend = f'''<div style="position:absolute;top:8px;left:8px;z-index:999;background:rgba(255,255,255,0.92);padding:8px 12px;border-radius:6px;font:11px sans-serif;box-shadow:0 2px 6px rgba(0,0,0,0.12);">
        <b>{title}</b><br><b>Types:</b> {type_legend}<br><b>Sectors:</b> {sect_legend}<br>
        <span style="color:#666;font-size:9px;">● Large=Sector ● Small=Overlap Group  Edge width=Overlap%</span></div>'''

    with tempfile.NamedTemporaryFile(mode="w", suffix=".html", delete=False) as f:
        net.save_graph(f.name)
        with open(f.name, "r") as rf:
            html = rf.read()
        html = html.replace("<body>", f"<body>{legend}")
        return html



def build_savings_sankey(df, selected_departments, dept_sector_map, short_name_map,
                         outside_split, dept_worktype_splits, dept_infra_splits, wt_categories, inf_categories,
                         total_wt_savings, total_inf_savings, total_ai_savings, latest_year):
    """Build a Sankey: Sector → Asset Type → Worktypes/InfraTypes/Labor → Savings/Optimal.
    Full budget flows from the beginning; savings and optimal split at the end."""
    labels = []
    label_idx = {}
    sources = []
    targets = []
    values = []
    colors = []

    def get_idx(name):
        if name not in label_idx:
            label_idx[name] = len(labels)
            labels.append(name)
        return label_idx[name]

    sector_colors_s = {
        "Public Safety": "rgba(139,41,66,0.6)", "Finance and Administration": "rgba(27,58,95,0.6)",
        "Infrastructure Services": "rgba(26,122,114,0.6)", "Community Services": "rgba(212,168,75,0.6)",
        "City Development": "rgba(217,119,6,0.6)", "Legislative and Election": "rgba(15,39,68,0.6)",
        "Regulatory": "rgba(59,130,246,0.6)",
    }
    asset_colors = {
        "Digital": "rgba(27,58,95,0.5)",
        "Physical": "rgba(26,122,114,0.5)",
        "Labor": "rgba(212,168,75,0.5)",
    }

    # Get all budgets for latest year
    all_mask = (
        (df["Department"].isin(selected_departments)) &
        (df["Year"] == latest_year)
    )
    dept_budgets = df[all_mask].groupby(["Department", "IT Type"])["Budget"].sum().reset_index()

    # Aggregate by sector
    sector_budgets = {}
    for dept in selected_departments:
        sector = dept_sector_map.get(dept, "Unknown")
        if sector not in sector_budgets:
            sector_budgets[sector] = {"digital": 0, "physical": 0, "labor": 0,
                                       "appdev": 0, "infra_b": 0, "sw": 0}
        db = dept_budgets[dept_budgets["Department"] == dept]
        os_val = db[db["IT Type"].isin(["IT Outside Services", "IT Outside Services*"])]["Budget"].sum()
        sw_val = db[db["IT Type"].str.strip() == "IT Software"]["Budget"].sum()
        hw_val = db[db["IT Type"].str.strip() == "IT Equipment"]["Budget"].sum()
        per_val = db[db["IT Type"].str.strip() == "IT Labor"]["Budget"].sum()

        appdev = os_val * outside_split["Application Development"]
        infra_b = os_val * outside_split["Infrastructure"]

        sector_budgets[sector]["digital"] += sw_val + appdev
        sector_budgets[sector]["appdev"] += appdev
        sector_budgets[sector]["sw"] += sw_val
        sector_budgets[sector]["physical"] += hw_val + infra_b
        sector_budgets[sector]["infra_b"] += infra_b
        sector_budgets[sector]["labor"] += per_val

    total_appdev = sum(s["appdev"] for s in sector_budgets.values())
    total_infra_b = sum(s["infra_b"] for s in sector_budgets.values())
    total_sw = sum(s["sw"] for s in sector_budgets.values())

    n_wt = len(wt_categories) if len(wt_categories) > 0 else 1
    ai_per_wt = total_ai_savings / n_wt

    # Accumulators for final split
    wt_savings_acc = {wt: 0.0 for wt in wt_categories}
    wt_budget_acc = {wt: 0.0 for wt in wt_categories}
    inf_savings_acc = {cat: 0.0 for cat in inf_categories}
    inf_budget_acc = {cat: 0.0 for cat in inf_categories}
    labor_total = 0.0

    for sector, data in sector_budgets.items():
        s_idx = get_idx(sector)

        # Sector → Digital
        if data["digital"] > 0:
            sources.append(s_idx); targets.append(get_idx("Digital"))
            values.append(data["digital"]); colors.append(asset_colors["Digital"])

        # Sector → Physical
        if data["physical"] > 0:
            sources.append(s_idx); targets.append(get_idx("Physical"))
            values.append(data["physical"]); colors.append(asset_colors["Physical"])

        # Sector → Labor
        if data["labor"] > 0:
            sources.append(s_idx); targets.append(get_idx("Labor"))
            values.append(data["labor"]); colors.append(asset_colors["Labor"])
            labor_total += data["labor"]

        # Savings proportions for this sector
        sec_wt_sav = (data["appdev"] / total_appdev) * total_wt_savings if total_appdev > 0 else 0
        sec_inf_sav = (data["infra_b"] / total_infra_b) * total_inf_savings if total_infra_b > 0 else 0
        sec_ai_weight = (data["sw"] / total_sw) if total_sw > 0 else 0

        sec_depts = [d for d in selected_departments if dept_sector_map.get(d) == sector]

        # Digital → Worktypes (full budget flows)
        for wt in wt_categories:
            avg_split = np.mean([dept_worktype_splits.get(d, {}).get(wt, 0.2) for d in sec_depts]) if sec_depts else 0.2
            wt_budget_val = data["appdev"] * avg_split + data["sw"] / n_wt
            if wt_budget_val > 0:
                sources.append(get_idx("Digital")); targets.append(get_idx(wt))
                values.append(wt_budget_val); colors.append("rgba(27,58,95,0.4)")
                wt_budget_acc[wt] += wt_budget_val
                wt_savings_acc[wt] += sec_wt_sav * avg_split + ai_per_wt * sec_ai_weight

        # Physical → Infra Types (full budget flows)
        for cat in inf_categories:
            avg_split = np.mean([dept_infra_splits.get(d, {}).get(cat, 0.25) for d in sec_depts]) if sec_depts else 0.25
            cat_budget_val = data["infra_b"] * avg_split
            hw_share = (data["physical"] - data["infra_b"]) * avg_split
            cat_total_val = cat_budget_val + hw_share
            if cat_total_val > 0:
                sources.append(get_idx("Physical")); targets.append(get_idx(cat))
                values.append(cat_total_val); colors.append("rgba(26,122,114,0.35)")
                inf_budget_acc[cat] += cat_total_val
                inf_savings_acc[cat] += sec_inf_sav * avg_split

    # Last level: Worktypes / Infra Types → Savings & Optimal
    savings_color = "rgba(5,150,105,0.5)"
    optimal_color = "rgba(71,85,105,0.3)"

    for wt in wt_categories:
        sav = wt_savings_acc[wt]
        budget = wt_budget_acc[wt]
        opt = max(0, budget - sav)
        if sav > 0:
            sources.append(get_idx(wt)); targets.append(get_idx("Savings"))
            values.append(sav); colors.append(savings_color)
        if opt > 0:
            sources.append(get_idx(wt)); targets.append(get_idx("Optimal"))
            values.append(opt); colors.append(optimal_color)

    for cat in inf_categories:
        sav = inf_savings_acc[cat]
        budget = inf_budget_acc[cat]
        opt = max(0, budget - sav)
        if sav > 0:
            sources.append(get_idx(cat)); targets.append(get_idx("Savings"))
            values.append(sav); colors.append(savings_color)
        if opt > 0:
            sources.append(get_idx(cat)); targets.append(get_idx("Optimal"))
            values.append(opt); colors.append(optimal_color)

    # Labor → Optimal (no savings)
    if labor_total > 0:
        sources.append(get_idx("Labor")); targets.append(get_idx("Optimal"))
        values.append(labor_total); colors.append(optimal_color)

    if not values:
        return None

    # Node colors
    node_colors = []
    for lbl in labels:
        if lbl in sector_colors_s:
            node_colors.append(sector_colors_s[lbl].replace("0.6", "0.9"))
        elif lbl in ["Digital", "Physical", "Labor"]:
            node_colors.append(asset_colors[lbl].replace("0.5", "0.9"))
        elif lbl in wt_categories:
            node_colors.append("rgba(27,58,95,0.9)")
        elif lbl in inf_categories:
            node_colors.append("rgba(26,122,114,0.9)")
        elif lbl == "Savings":
            node_colors.append("rgba(5,150,105,0.9)")
        elif lbl == "Optimal":
            node_colors.append("rgba(71,85,105,0.9)")
        else:
            node_colors.append("rgba(150,150,150,0.9)")

    # Format labels with values
    node_values = [0.0] * len(labels)
    for s, t, v in zip(sources, targets, values):
        node_values[s] += v  # not perfect but gives sense of throughput

    display_labels = []
    for i, lbl in enumerate(labels):
        display_labels.append(lbl)

    fig = go.Figure(go.Sankey(
        node=dict(pad=20, thickness=25, line=dict(color="black", width=0.5),
                  label=display_labels, color=node_colors),
        link=dict(source=sources, target=targets, value=values, color=colors),
    ))
    fig.update_layout(margin=dict(l=10, r=10, t=30, b=10), height=650,
                      font=dict(size=13, color="#000000"))
    return fig



# ── Sidebar filters ──
st.sidebar.image("https://upload.wikimedia.org/wikipedia/commons/thumb/9/9b/Flag_of_Chicago%2C_Illinois.svg/200px-Flag_of_Chicago%2C_Illinois.svg.png", width=120)
st.sidebar.title("🔍 Filters")

all_years = sorted(df["Year"].unique())
selected_years = st.sidebar.select_slider(
    "Year Range",
    options=all_years,
    value=(all_years[0], all_years[-1])
)

all_sectors = sorted(df["Sector"].unique())
selected_sector = st.sidebar.selectbox(
    "Sector",
    options=["All Sectors"] + all_sectors,
    index=0,
)

if selected_sector == "All Sectors":
    sector_departments = sorted(df["Department"].unique())
else:
    sector_departments = sorted(df[df["Sector"] == selected_sector]["Department"].unique())

selected_department = st.sidebar.selectbox(
    "Department",
    options=["All Departments"] + sector_departments,
    index=0,
)

if selected_department == "All Departments":
    selected_departments = sector_departments
else:
    selected_departments = [selected_department]

all_it_types = sorted(df["IT Type"].unique())
selected_it_type = st.sidebar.selectbox(
    "IT Spend Types",
    options=["All Types"] + all_it_types,
    index=0,
)

if selected_it_type == "All Types":
    selected_it_types = all_it_types
else:
    selected_it_types = [selected_it_type]

st.sidebar.divider()
with st.sidebar.expander("🔮 Projection Settings", expanded=False):
    growth_multiplier = st.slider(
        "Growth Rate Multiplier",
        min_value=0.5,
        max_value=1.5,
        value=1.0,
        step=0.05,
        help="1.0 = trend as-is, 0.5 = half the growth, 1.5 = 50% faster growth"
    )
    projection_noise = st.slider(
        "Projection Noise %",
        min_value=0,
        max_value=20,
        value=7,
        step=1,
        help="0 = smooth lines, 5-10 = realistic variation, 20 = high volatility"
    ) / 100

with st.sidebar.expander("🧮 Worktype Distribution", expanded=False):
    st.caption("Infrastructure + App Development = 100%")
    pct_infra = st.slider("Infrastructure %", 0, 100, 35, 5, key="pct_infra")
    pct_appdev = 100 - pct_infra
    st.info(f"Application Development: **{pct_appdev}%**")

outside_split = {
    "Infrastructure": pct_infra / 100,
    "Application Development": pct_appdev / 100,
}

@st.cache_data
def load_worktype_splits():
    import os
    script_dir = os.path.dirname(os.path.abspath(__file__))
    csv_path = os.path.join(script_dir, "worktype_splits.csv")
    return pd.read_csv(csv_path)

df_wt_splits = load_worktype_splits()

# Build department-level worktype split lookup: {dept: {wt: pct}}
wt_categories = ["Workflow", "Analysis", "Collaboration", "Transactions", "External Facing"]
dept_worktype_splits = {}
for _, row in df_wt_splits.iterrows():
    dept_worktype_splits[row["Department"]] = {
        wt: row[wt] / 100 for wt in wt_categories
    }

# For backward compatibility, compute a weighted average worktype split for selected departments
# (used in KPIs and aggregate views)
worktype_split = {wt: 0.0 for wt in wt_categories}
if selected_departments:
    for wt in wt_categories:
        worktype_split[wt] = np.mean([
            dept_worktype_splits.get(d, {}).get(wt, 0.2)
            for d in selected_departments
        ])

pct_wt_total = 100  # Always valid since CSV is pre-validated

with st.sidebar.expander("📈 Worktype Growth Rates", expanded=False):
    st.caption("Independent growth multiplier per worktype (applied on top of the global growth multiplier).")
    wt_growth = {}
    wt_defaults = {
        "Workflow": 0.90,
        "Analysis": 1.30,
        "Collaboration": 0.85,
        "Transactions": 1.05,
        "External Facing": 1.45,
    }
    for wt, default in wt_defaults.items():
        wt_growth[wt] = st.slider(
            f"{wt}", 0.5, 2.0, default, 0.05, key=f"wt_growth_{wt}"
        )

# Build department-level infrastructure split lookup: {dept: {cat: pct}}
inf_categories = ["Integration", "Access", "Data", "Communications"]
dept_infra_splits = {}
for _, row in df_wt_splits.iterrows():
    dept_infra_splits[row["Department"]] = {
        cat: row[cat] / 100 for cat in inf_categories
    }

# Weighted average infra split for selected departments
infra_split = {cat: 0.0 for cat in inf_categories}
if selected_departments:
    for cat in inf_categories:
        infra_split[cat] = np.mean([
            dept_infra_splits.get(d, {}).get(cat, 0.25)
            for d in selected_departments
        ])

pct_infra_total = 100  # Always valid since CSV is pre-validated

with st.sidebar.expander("📈 Infrastructure Growth Rates", expanded=False):
    st.caption("Independent growth multiplier per infrastructure category.")
    infra_growth = {}
    infra_defaults = {
        "Integration": 1.20,
        "Access": 0.95,
        "Data": 1.35,
        "Communications": 0.80,
    }
    for cat, default in infra_defaults.items():
        infra_growth[cat] = st.slider(
            f"{cat}", 0.5, 2.0, default, 0.05, key=f"infra_growth_{cat}"
        )

with st.sidebar.expander("💰 Savings Factors", expanded=False):
    st.caption("Control how much of identified overlap/fragmentation translates to actual savings.")
    wt_savings_factor = st.slider("Worktype Overlap Savings %", 0, 100, 50, 5, key="wt_savings") / 100
    inf_savings_factor = st.slider("Infrastructure Fragmentation Savings %", 0, 100, 50, 5, key="inf_savings") / 100
    ai_replacement_factor = st.slider("AI License Replacement %", 0, 100, 50, 5, key="ai_replace") / 100

# ── Load overlap & fragmentation data ──
@st.cache_data
def load_overlap_data():
    import os
    script_dir = os.path.dirname(os.path.abspath(__file__))
    df_wt_overlap = pd.read_excel(
        os.path.join(script_dir, "Worktype_Overlaping.xlsx"),
        sheet_name="Application_Overlap_Grouped"
    )
    df_inf_frag = pd.read_excel(
        os.path.join(script_dir, "Infrastructure_Fragmentation.xlsx"),
        sheet_name="Infrastructure_Overlap_Grouped"
    )
    return df_wt_overlap, df_inf_frag

df_wt_overlap, df_inf_frag = load_overlap_data()

# ── Load SaaS license data ──
@st.cache_data
def load_saas_data():
    import os
    script_dir = os.path.dirname(os.path.abspath(__file__))
    df_saas = pd.read_excel(
        os.path.join(script_dir, "SaaSvsIA.xlsx"),
        sheet_name="Per_Person_Licenses_With_AI_Rea"
    )
    df_saas.columns = ["License", "Category", "Price Per User", "AI Replaceability", "Replaceability Reason"]
    return df_saas

df_saas = load_saas_data()

st.sidebar.divider()
if st.sidebar.button("🔄 Reset All Filters", use_container_width=True):
    st.cache_data.clear()
    for key in list(st.session_state.keys()):
        del st.session_state[key]
    st.rerun()

# ── Filter data ──
mask = (
    (df["Year"] >= selected_years[0]) &
    (df["Year"] <= selected_years[1]) &
    (df["Department"].isin(selected_departments)) &
    (df["IT Type"].isin(selected_it_types))
)
filtered = df[mask]

# ── Title ──
st.title("🏙️ City of Chicago — Digital Asset Model")
st.markdown(f"Analyzing **{len(selected_departments)}** departments across **{selected_years[0]}–{selected_years[1]}**")

# ── Tabs ──
tab_overview, tab_projections, tab_outside, tab_worktypes, tab_infra, tab_savings, tab_lifecycle = st.tabs([
    "📊 Budget Overview",
    "🔮 Projections",
    "🧮 Outside Services",
    "💻 Applications",
    "🏗️ Infrastructure",
    "💰 Savings",
    "♻️ Lifecycle Optimization",
])

latest_year = selected_years[1]
prev_year = latest_year - 1

# ══════════════════════════════════════════════════════════════
# TAB 1: BUDGET OVERVIEW
# ══════════════════════════════════════════════════════════════
with tab_overview:

    # ── KPI row ──
    total_budget = filtered["Budget"].sum()
    budget_latest = filtered[filtered["Year"] == latest_year]["Budget"].sum()
    budget_prev = filtered[filtered["Year"] == prev_year]["Budget"].sum()
    yoy_change = ((budget_latest - budget_prev) / budget_prev * 100) if budget_prev else 0
    avg_per_dept = budget_latest / len(selected_departments) if selected_departments else 0
    top_dept = (
        filtered[filtered["Year"] == latest_year]
        .groupby("Department")["Budget"].sum()
        .idxmax() if not filtered.empty else "N/A"
    )

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Total Budget (All Selected Years)", fmt(total_budget))
    col2.metric(f"Budget {latest_year}", fmt(budget_latest), f"{yoy_change:+.1f}% vs {prev_year}")
    col3.metric(f"Avg per Dept ({latest_year})", fmt(avg_per_dept))
    col4.metric(f"Top Dept ({latest_year})", top_dept[:30])

    st.divider()

    # ── Trend + Breakdown ──
    r1c1, r1c2 = st.columns([3, 2])

    with r1c1:
        st.subheader("📈 Budget Trend by Year")
        trend = filtered.groupby("Year")["Budget"].sum().reset_index()
        fig_trend = px.area(
            trend, x="Year", y="Budget",
            labels={"Budget": "Total Budget ($)"},
            color_discrete_sequence=["#1B3A5F"]
        )
        fig_trend.update_layout(
            yaxis_tickprefix="$", yaxis_tickformat=",",
            xaxis=dict(dtick=1),
            margin=dict(l=20, r=20, t=10, b=20),
            height=370,
        )
        st.plotly_chart(fig_trend, use_container_width=True)

    with r1c2:
        st.subheader("🧩 Spend by IT Type")
        by_type = filtered.groupby("IT Type")["Budget"].sum().reset_index()
        type_colors_map = {"IT Equipment": "#1B3A5F", "IT Labor": "#1A7A72", "IT Outside Services": "#D4A84B",
                           "IT Outside Services*": "#8B2942", "IT Software": "#3B82F6"}
        fig_pie = px.pie(
            by_type, names="IT Type", values="Budget",
            color="IT Type", color_discrete_map=type_colors_map,
            hole=0.45
        )
        fig_pie.update_traces(
            textinfo="percent+label+value", textposition="outside",
            texttemplate="%{label}<br>%{percent}<br>$%{value:,.0f}"
        )
        fig_pie.update_layout(
            margin=dict(l=20, r=20, t=10, b=20),
            height=370,
            showlegend=False
        )
        st.plotly_chart(fig_pie, use_container_width=True)

    # ── Top departments + type trend ──
    r2c1, r2c2 = st.columns(2)

    with r2c1:
        st.subheader("🏢 Top 6 Departments by Budget (2021–2026)")
        dept_budget = (
            filtered
            .groupby("Department")["Budget"].sum()
            .sort_values(ascending=True)
            .tail(6)
            .reset_index()
        )
        dept_budget["Short Name"] = dept_budget["Department"].map(short_name_map)
        fig_bar = px.bar(
            dept_budget, x="Budget", y="Short Name",
            orientation="h",
            color="Budget",
            color_continuous_scale=[[0,"#F8FAFC"],[0.5,"#1A7A72"],[1,"#0F2744"]],
        )
        fig_bar.update_layout(
            xaxis_tickprefix="$", xaxis_tickformat=",",
            margin=dict(l=20, r=20, t=10, b=20),
            height=450,
            coloraxis_showscale=False,
            yaxis_title=""
        )
        st.plotly_chart(fig_bar, use_container_width=True)

    with r2c2:
        st.subheader("📊 Budget Trend by IT Type")
        type_trend = filtered.groupby(["Year", "IT Type"])["Budget"].sum().reset_index()
        fig_type = px.bar(
            type_trend, x="Year", y="Budget", color="IT Type",
            barmode="stack",
            color_discrete_sequence=["#1B3A5F", "#1A7A72", "#D4A84B", "#8B2942", "#7C3AED"],
        )
        fig_type.update_layout(
            yaxis_tickprefix="$", yaxis_tickformat=",",
            xaxis=dict(dtick=1),
            margin=dict(l=20, r=20, t=10, b=20),
            height=450,
            legend=dict(orientation="h", yanchor="bottom", y=-0.25),
        )
        st.plotly_chart(fig_type, use_container_width=True)

    # ── Heatmap ──
    st.subheader("🔥 Department × Year Heatmap")
    heatmap_data = (
        filtered.groupby(["Department", "Year"])["Budget"].sum()
        .unstack(fill_value=0)
    )
    top_20 = heatmap_data.sum(axis=1).sort_values(ascending=False).head(20).index
    heatmap_data = heatmap_data.loc[top_20].sort_values(by=heatmap_data.columns[-1], ascending=True)
    heatmap_data.index = [short_name_map.get(d, d) for d in heatmap_data.index]

    fig_heat = px.imshow(
        heatmap_data,
        labels=dict(x="Year", y="Department", color="Budget ($)"),
        color_continuous_scale=[[0,"#F8FAFC"],[0.5,"#1A7A72"],[1,"#0F2744"]],
        aspect="auto",
    )
    format_fn = lambda v: f"${v/1e6:.1f}M" if v >= 1e6 else f"${v/1e3:.0f}K"
    try:
        text_labels = heatmap_data.map(format_fn).values
    except AttributeError:
        text_labels = heatmap_data.applymap(format_fn).values
    fig_heat.update_traces(text=text_labels, texttemplate="%{text}")
    fig_heat.update_layout(
        margin=dict(l=20, r=20, t=10, b=20),
        height=550,
    )
    st.plotly_chart(fig_heat, use_container_width=True)

    # ── Data table ──
    with st.expander("📋 View Raw Data"):
        st.dataframe(
            filtered.sort_values(["Year", "Department"]),
            use_container_width=True,
            height=400
        )

# ══════════════════════════════════════════════════════════════
# TAB 2: PROJECTIONS
# ══════════════════════════════════════════════════════════════
with tab_projections:

    st.caption(f"Projections based on linear regression of 2021–2026 data, anchored from last actual value. Growth multiplier: **{growth_multiplier}x**")

    proj_mask = (
        (df["Department"].isin(selected_departments)) &
        (df["IT Type"].isin(selected_it_types))
    )
    proj_data = df[proj_mask]

    p1c1, p1c2 = st.columns(2)

    # ── Projection by Department ──
    with p1c1:
        st.subheader("📈 By Department")
        dept_hist = proj_data.groupby(["Department", "Year"])["Budget"].sum().unstack(fill_value=0)
        proj_rows = []
        for dept in dept_hist.index:
            hist_vals = [dept_hist.loc[dept, y] if y in dept_hist.columns else 0 for y in all_hist_years]
            proj_vals = project_series(all_hist_years, hist_vals, future_years, growth_multiplier, projection_noise, seed=hash(dept) % 10000)
            for y, v in zip(all_hist_years, hist_vals):
                proj_rows.append({"Department": dept, "Year": y, "Budget": v, "Type": "Actual"})
            for y, v in zip(future_years, proj_vals):
                proj_rows.append({"Department": dept, "Year": y, "Budget": v, "Type": "Projected"})
        df_dept_proj = pd.DataFrame(proj_rows)

        total_proj = df_dept_proj.groupby(["Year", "Type"])["Budget"].sum().reset_index()
        fig_proj_total = go.Figure()
        actual = total_proj[total_proj["Type"] == "Actual"]
        projected = total_proj[total_proj["Type"] == "Projected"]
        last_actual = actual[actual["Year"] == actual["Year"].max()]
        projected_bridge = pd.concat([last_actual.assign(Type="Projected"), projected])

        fig_proj_total.add_trace(go.Scatter(
            x=actual["Year"], y=actual["Budget"],
            mode="lines+markers", name="Actual",
            line=dict(color="#1B3A5F", width=3), marker=dict(size=8)
        ))
        fig_proj_total.add_trace(go.Scatter(
            x=projected_bridge["Year"], y=projected_bridge["Budget"],
            mode="lines+markers", name="Projected",
            line=dict(color="#D4A84B", width=3, dash="dash"), marker=dict(size=8, symbol="diamond")
        ))
        fig_proj_total.update_layout(
            yaxis_tickprefix="$", yaxis_tickformat=",",
            xaxis=dict(dtick=1),
            margin=dict(l=20, r=20, t=10, b=20), height=400,
            legend=dict(orientation="h", yanchor="bottom", y=-0.2),
        )
        st.plotly_chart(fig_proj_total, use_container_width=True)

        dept_proj_summary = df_dept_proj[df_dept_proj["Type"] == "Projected"].pivot_table(
            index="Department", columns="Year", values="Budget", aggfunc="sum"
        )

    # ── Projection by IT Type ──
    with p1c2:
        st.subheader("📈 By IT Spend Type")
        type_hist = proj_data.groupby(["IT Type", "Year"])["Budget"].sum().unstack(fill_value=0)
        proj_type_rows = []
        for it_type in type_hist.index:
            hist_vals = [type_hist.loc[it_type, y] if y in type_hist.columns else 0 for y in all_hist_years]
            proj_vals = project_series(all_hist_years, hist_vals, future_years, growth_multiplier, projection_noise, seed=hash(it_type) % 10000)
            for y, v in zip(all_hist_years, hist_vals):
                proj_type_rows.append({"IT Type": it_type, "Year": y, "Budget": v, "Type": "Actual"})
            for y, v in zip(future_years, proj_vals):
                proj_type_rows.append({"IT Type": it_type, "Year": y, "Budget": v, "Type": "Projected"})
        df_type_proj = pd.DataFrame(proj_type_rows)

        colors = {"IT Equipment": "#1B3A5F", "IT Labor": "#1A7A72", "IT Outside Services": "#D4A84B",
                  "IT Outside Services*": "#8B2942", "IT Software": "#3B82F6"}
        fig_proj_type = go.Figure()
        for it_type in sorted(df_type_proj["IT Type"].unique()):
            subset = df_type_proj[df_type_proj["IT Type"] == it_type]
            actual_t = subset[subset["Type"] == "Actual"].sort_values("Year")
            proj_t = subset[subset["Type"] == "Projected"].sort_values("Year")
            last_a = actual_t[actual_t["Year"] == actual_t["Year"].max()]
            proj_bridge = pd.concat([last_a, proj_t])
            color = colors.get(it_type, "#999999")
            fig_proj_type.add_trace(go.Scatter(
                x=actual_t["Year"], y=actual_t["Budget"],
                mode="lines+markers", name=it_type,
                legendgroup=it_type,
                line=dict(color=color, width=2), marker=dict(size=6),
            ))
            fig_proj_type.add_trace(go.Scatter(
                x=proj_bridge["Year"], y=proj_bridge["Budget"],
                mode="lines+markers", name=it_type,
                legendgroup=it_type, showlegend=False,
                line=dict(color=color, width=2, dash="dash"),
                marker=dict(size=6, symbol="diamond"),
            ))

        fig_proj_type.update_layout(
            yaxis_tickprefix="$", yaxis_tickformat=",",
            xaxis=dict(dtick=1),
            margin=dict(l=20, r=20, t=10, b=20), height=400,
            legend=dict(orientation="h", yanchor="bottom", y=-0.35, font=dict(size=10)),
        )
        st.plotly_chart(fig_proj_type, use_container_width=True)

    st.divider()

    # ── Growth Rates ──
    st.subheader("📊 CAGR Growth Rates (2021–2026)")

    # Compute CAGR helper
    def compute_cagr(start_val, end_val, years):
        if start_val > 0 and end_val > 0 and years > 0:
            return (end_val / start_val) ** (1 / years) - 1
        return 0

    first_year = all_hist_years[0]
    last_hist_year = all_hist_years[-1]
    n_years = last_hist_year - first_year

    gr_c1, gr_c2, gr_c3 = st.columns(3)

    with gr_c1:
        st.subheader("By Sector")
        sector_growth = []
        for sector in sorted(df["Sector"].unique()):
            sec_depts = df[df["Sector"] == sector]["Department"].unique()
            sec_mask = (df["Department"].isin(sec_depts)) & (df["Department"].isin(selected_departments)) & (df["IT Type"].isin(selected_it_types))
            start = df[sec_mask & (df["Year"] == first_year)]["Budget"].sum()
            end = df[sec_mask & (df["Year"] == last_hist_year)]["Budget"].sum()
            cagr = compute_cagr(start, end, n_years)
            if start > 0 or end > 0:
                sector_growth.append({"Sector": sector, "CAGR": cagr * 100})
        if sector_growth:
            df_sg = pd.DataFrame(sector_growth).sort_values("CAGR", ascending=True)
            fig_sg = px.bar(df_sg, x="CAGR", y="Sector", orientation="h",
                            color="CAGR", color_continuous_scale=[[0,"#DC2626"],[0.5,"#D4A84B"],[1,"#059669"]], color_continuous_midpoint=0)
            fig_sg.update_layout(xaxis_ticksuffix="%", margin=dict(l=20, r=20, t=10, b=20),
                                 height=300, coloraxis_showscale=False, yaxis_title="")
            st.plotly_chart(fig_sg, use_container_width=True)

    with gr_c2:
        st.subheader("By Department (Top 10)")
        dept_growth = []
        for dept in selected_departments:
            d_mask = (df["Department"] == dept) & (df["IT Type"].isin(selected_it_types))
            start = df[d_mask & (df["Year"] == first_year)]["Budget"].sum()
            end = df[d_mask & (df["Year"] == last_hist_year)]["Budget"].sum()
            cagr = compute_cagr(start, end, n_years)
            if start > 0 or end > 0:
                dept_growth.append({"Department": short_name_map.get(dept, dept), "CAGR": cagr * 100})
        if dept_growth:
            df_dg = pd.DataFrame(dept_growth).sort_values("CAGR", ascending=False).head(10).sort_values("CAGR", ascending=True)
            fig_dg = px.bar(df_dg, x="CAGR", y="Department", orientation="h",
                            color="CAGR", color_continuous_scale=[[0,"#DC2626"],[0.5,"#D4A84B"],[1,"#059669"]], color_continuous_midpoint=0)
            fig_dg.update_layout(xaxis_ticksuffix="%", margin=dict(l=20, r=20, t=10, b=20),
                                 height=350, coloraxis_showscale=False, yaxis_title="")
            st.plotly_chart(fig_dg, use_container_width=True)

    with gr_c3:
        st.subheader("By IT Type")
        type_growth = []
        for it_type in sorted(df["IT Type"].unique()):
            t_mask = (df["Department"].isin(selected_departments)) & (df["IT Type"] == it_type)
            start = df[t_mask & (df["Year"] == first_year)]["Budget"].sum()
            end = df[t_mask & (df["Year"] == last_hist_year)]["Budget"].sum()
            cagr = compute_cagr(start, end, n_years)
            if start > 0 or end > 0:
                type_growth.append({"IT Type": it_type, "CAGR": cagr * 100})
        if type_growth:
            df_tg = pd.DataFrame(type_growth).sort_values("CAGR", ascending=True)
            fig_tg = px.bar(df_tg, x="CAGR", y="IT Type", orientation="h",
                            color="CAGR", color_continuous_scale=[[0,"#DC2626"],[0.5,"#D4A84B"],[1,"#059669"]], color_continuous_midpoint=0)
            fig_tg.update_layout(xaxis_ticksuffix="%", margin=dict(l=20, r=20, t=10, b=20),
                                 height=300, coloraxis_showscale=False, yaxis_title="")
            st.plotly_chart(fig_tg, use_container_width=True)

# ══════════════════════════════════════════════════════════════
# TAB 3: OUTSIDE SERVICES SPLIT
# ══════════════════════════════════════════════════════════════
with tab_outside:

    st.caption(f"Combined IT Outside Services & IT Outside Services* split into **Infrastructure ({pct_infra}%)** and **Application Development ({pct_appdev}%)**. Adjust in sidebar.")

    # Combine both Outside Services types
    os_mask = (
        (df["Department"].isin(selected_departments)) &
        (df["IT Type"].isin(["IT Outside Services", "IT Outside Services*"]))
    )
    os_data = df[os_mask]

    # ── KPIs ──
    os_total = os_data["Budget"].sum()
    os_latest = os_data[os_data["Year"] == latest_year]["Budget"].sum()
    os_prev = os_data[os_data["Year"] == prev_year]["Budget"].sum()
    os_yoy = ((os_latest - os_prev) / os_prev * 100) if os_prev else 0

    kc1, kc2, kc3, kc4 = st.columns(4)
    kc1.metric("Outside Services Total", fmt(os_total))
    kc2.metric(f"Outside Services {latest_year}", fmt(os_latest), f"{os_yoy:+.1f}% vs {prev_year}")
    kc3.metric(f"Infrastructure ({latest_year})", fmt(os_latest * outside_split['Infrastructure']))
    kc4.metric(f"App Development ({latest_year})", fmt(os_latest * outside_split['Application Development']))

    st.divider()

    # Build actual + projected subcategory data (with slight random variation per year)
    os_by_year = os_data.groupby("Year")["Budget"].sum().reset_index()
    os_sub_rows = []
    rng_split = np.random.default_rng(42)
    base_infra_pct = outside_split["Infrastructure"]
    for _, row in os_by_year.iterrows():
        noise = rng_split.uniform(-0.05, 0.05)
        yr_infra_pct = max(0.1, min(0.9, base_infra_pct + noise))
        yr_appdev_pct = 1.0 - yr_infra_pct
        os_sub_rows.append({
            "Year": row["Year"], "Subcategory": "Infrastructure",
            "Budget": row["Budget"] * yr_infra_pct, "Period": "Actual",
        })
        os_sub_rows.append({
            "Year": row["Year"], "Subcategory": "Application Development",
            "Budget": row["Budget"] * yr_appdev_pct, "Period": "Actual",
        })

    os_hist_vals = [os_by_year[os_by_year["Year"] == y]["Budget"].sum() for y in all_hist_years]
    os_proj_vals = project_series(all_hist_years, os_hist_vals, future_years, growth_multiplier, projection_noise, seed=100)
    for fy, fv in zip(future_years, os_proj_vals):
        noise = rng_split.uniform(-0.05, 0.05)
        yr_infra_pct = max(0.1, min(0.9, base_infra_pct + noise))
        yr_appdev_pct = 1.0 - yr_infra_pct
        os_sub_rows.append({
            "Year": fy, "Subcategory": "Infrastructure",
            "Budget": fv * yr_infra_pct, "Period": "Projected",
        })
        os_sub_rows.append({
            "Year": fy, "Subcategory": "Application Development",
            "Budget": fv * yr_appdev_pct, "Period": "Projected",
        })
    df_os_sub = pd.DataFrame(os_sub_rows)

    sub_colors = {"Infrastructure": "#1B3A5F", "Application Development": "#D4A84B"}

    os_c1, os_c2 = st.columns(2)

    with os_c1:
        st.subheader("📊 Infrastructure vs App Development by Year")
        fig_os_stack = go.Figure()
        for subcat in outside_split:
            subset = df_os_sub[df_os_sub["Subcategory"] == subcat].sort_values("Year")
            actual_b = subset[subset["Period"] == "Actual"]
            proj_b = subset[subset["Period"] == "Projected"]
            color = sub_colors[subcat]
            fig_os_stack.add_trace(go.Bar(
                x=actual_b["Year"], y=actual_b["Budget"],
                name=subcat, legendgroup=subcat,
                marker_color=color,
            ))
            fig_os_stack.add_trace(go.Bar(
                x=proj_b["Year"], y=proj_b["Budget"],
                name=subcat, legendgroup=subcat, showlegend=False,
                marker_color=color, marker_pattern_shape="/",
            ))
        fig_os_stack.update_layout(
            barmode="stack",
            yaxis_tickprefix="$", yaxis_tickformat=",",
            xaxis=dict(dtick=1),
            margin=dict(l=20, r=20, t=10, b=20), height=450,
            legend=dict(orientation="h", yanchor="bottom", y=-0.25),
        )
        st.plotly_chart(fig_os_stack, use_container_width=True)

    with os_c2:
        st.subheader(f"🍩 Allocation Split ({latest_year})")
        pie_data = df_os_sub[(df_os_sub["Year"] == latest_year) & (df_os_sub["Period"] == "Actual")]
        fig_os_pie = px.pie(
            pie_data, names="Subcategory", values="Budget",
            color="Subcategory", color_discrete_map=sub_colors, hole=0.45,
        )
        fig_os_pie.update_traces(
            textinfo="percent+label+value", textposition="outside",
            texttemplate="%{label}<br>%{percent}<br>$%{value:,.0f}"
        )
        fig_os_pie.update_layout(
            margin=dict(l=20, r=20, t=10, b=20), height=450, showlegend=False,
        )
        st.plotly_chart(fig_os_pie, use_container_width=True)

    # ── Trend lines ──
    st.subheader("📈 Subcategory Trends (Actual + Projected)")
    fig_os_trend = go.Figure()
    for subcat in outside_split:
        subset = df_os_sub[df_os_sub["Subcategory"] == subcat]
        actual_s = subset[subset["Period"] == "Actual"].sort_values("Year")
        proj_s = subset[subset["Period"] == "Projected"].sort_values("Year")
        last_a = actual_s[actual_s["Year"] == actual_s["Year"].max()]
        proj_bridge = pd.concat([last_a.assign(Period="Projected"), proj_s])
        color = sub_colors[subcat]
        fig_os_trend.add_trace(go.Scatter(
            x=actual_s["Year"], y=actual_s["Budget"],
            mode="lines+markers", name=subcat,
            legendgroup=subcat,
            line=dict(color=color, width=3), marker=dict(size=8),
        ))
        fig_os_trend.add_trace(go.Scatter(
            x=proj_bridge["Year"], y=proj_bridge["Budget"],
            mode="lines+markers", name=subcat,
            legendgroup=subcat, showlegend=False,
            line=dict(color=color, width=3, dash="dash"), marker=dict(size=8, symbol="diamond"),
        ))
    fig_os_trend.update_layout(
        yaxis_tickprefix="$", yaxis_tickformat=",",
        xaxis=dict(dtick=1),
        margin=dict(l=20, r=20, t=10, b=20), height=400,
        legend=dict(orientation="h", yanchor="bottom", y=-0.25),
    )
    st.plotly_chart(fig_os_trend, use_container_width=True)

    # ── Summary table ──
    st.subheader("📋 Summary Table")
    os_summary = df_os_sub.pivot_table(
        index="Subcategory", columns="Year", values="Budget", aggfunc="sum"
    )
    os_summary.columns = [str(int(c)) for c in os_summary.columns]
    os_summary["Total"] = os_summary.sum(axis=1)
    st.dataframe(os_summary.style.format("${:,.0f}"), use_container_width=True, height=150)

    # ── Department breakdown ──
    st.subheader(f"🏢 Outside Services by Department — Top 10 ({latest_year})")
    os_dept_all = os_data[os_data["Year"] == latest_year].groupby("Department")["Budget"].sum().sort_values(ascending=False).reset_index()
    if len(os_dept_all) > 10:
        top10 = os_dept_all.head(10)
        others_val = os_dept_all.iloc[10:]["Budget"].sum()
        others_row = pd.DataFrame([{"Department": "Others", "Budget": others_val}])
        os_dept = pd.concat([top10, others_row], ignore_index=True).sort_values("Budget", ascending=True)
    else:
        os_dept = os_dept_all.sort_values("Budget", ascending=True)
    os_dept["Short Name"] = os_dept["Department"].map(lambda d: short_name_map.get(d, d))
    os_dept["Infrastructure"] = os_dept["Budget"] * outside_split["Infrastructure"]
    os_dept["App Development"] = os_dept["Budget"] * outside_split["Application Development"]

    fig_os_dept = go.Figure()
    fig_os_dept.add_trace(go.Bar(
        y=os_dept["Short Name"], x=os_dept["Infrastructure"],
        name="Infrastructure", orientation="h", marker_color="#1B3A5F",
    ))
    fig_os_dept.add_trace(go.Bar(
        y=os_dept["Short Name"], x=os_dept["App Development"],
        name="App Development", orientation="h", marker_color="#D4A84B",
    ))
    fig_os_dept.update_layout(
        barmode="stack",
        xaxis_tickprefix="$", xaxis_tickformat=",",
        margin=dict(l=20, r=20, t=10, b=20),
        height=max(400, len(os_dept) * 25),
        legend=dict(orientation="h", yanchor="bottom", y=-0.15),
        yaxis_title="",
    )
    st.plotly_chart(fig_os_dept, use_container_width=True)

# ══════════════════════════════════════════════════════════════
# TAB 4: APP DEV WORKTYPES
# ══════════════════════════════════════════════════════════════
with tab_worktypes:

    st.caption(
        f"Application Development budget ({pct_appdev}% of Outside Services) split across 5 worktypes per department. "
        f"Splits loaded from **worktype_splits.csv**. Growth multiplier: **{growth_multiplier}x**."
    )

    # Get Outside Services by department and year
    os_mask_wt = (
        (df["Department"].isin(selected_departments)) &
        (df["IT Type"].isin(["IT Outside Services", "IT Outside Services*"]))
    )
    os_data_wt = df[os_mask_wt]

    # Build worktype data using department-level splits
    wt_rows = []
    for dept in selected_departments:
        dept_os = os_data_wt[os_data_wt["Department"] == dept].groupby("Year")["Budget"].sum().reset_index()
        dept_os["Budget"] = dept_os["Budget"] * outside_split["Application Development"]
        splits = dept_worktype_splits.get(dept, {wt: 0.2 for wt in wt_categories})

        for _, row in dept_os.iterrows():
            for wt in wt_categories:
                wt_rows.append({
                    "Year": row["Year"], "Worktype": wt,
                    "Budget": row["Budget"] * splits[wt], "Period": "Actual",
                    "Department": dept,
                })

        # Project each worktype independently per department
        for wt in wt_categories:
            wt_hist = [
                dept_os[dept_os["Year"] == y]["Budget"].sum() * splits[wt]
                for y in all_hist_years
            ]
            compound_multiplier = growth_multiplier * wt_growth[wt]
            wt_proj = project_series(all_hist_years, wt_hist, future_years, compound_multiplier, projection_noise, seed=hash((dept, wt)) % 10000)
            for fy, fv in zip(future_years, wt_proj):
                wt_rows.append({
                    "Year": fy, "Worktype": wt,
                    "Budget": fv, "Period": "Projected",
                    "Department": dept,
                })

    df_wt = pd.DataFrame(wt_rows)

    # Aggregate for charts
    df_wt_agg = df_wt.groupby(["Year", "Worktype", "Period"])["Budget"].sum().reset_index()

    # ── KPIs ──
    ad_latest_by_wt = df_wt_agg[(df_wt_agg["Year"] == latest_year) & (df_wt_agg["Period"] == "Actual")]
    wk1, wk2, wk3, wk4, wk5 = st.columns(5)
    for col, wt in zip([wk1, wk2, wk3, wk4, wk5], wt_categories):
        val = ad_latest_by_wt[ad_latest_by_wt["Worktype"] == wt]["Budget"].sum()
        col.metric(wt, fmt(val), f"{wt_growth[wt]}x growth")

    st.divider()

    wt_colors = {
        "Workflow": "#1B3A5F",
        "Analysis": "#1A7A72",
        "Collaboration": "#D4A84B",
        "Transactions": "#8B2942",
        "External Facing": "#7C3AED",
    }

    wt_c1, wt_c2 = st.columns(2)

    with wt_c1:
        st.subheader("📊 Worktypes by Year")
        fig_wt_stack = go.Figure()
        for wt in wt_categories:
            subset = df_wt_agg[df_wt_agg["Worktype"] == wt].sort_values("Year")
            actual_b = subset[subset["Period"] == "Actual"]
            proj_b = subset[subset["Period"] == "Projected"]
            color = wt_colors[wt]
            fig_wt_stack.add_trace(go.Bar(
                x=actual_b["Year"], y=actual_b["Budget"],
                name=wt, legendgroup=wt,
                marker_color=color,
            ))
            fig_wt_stack.add_trace(go.Bar(
                x=proj_b["Year"], y=proj_b["Budget"],
                name=wt, legendgroup=wt, showlegend=False,
                marker_color=color, marker_pattern_shape="/",
            ))
        fig_wt_stack.update_layout(
            barmode="stack",
            yaxis_tickprefix="$", yaxis_tickformat=",",
            xaxis=dict(dtick=1),
            margin=dict(l=20, r=20, t=10, b=20), height=450,
            legend=dict(orientation="h", yanchor="bottom", y=-0.3, font=dict(size=10)),
        )
        st.plotly_chart(fig_wt_stack, use_container_width=True)

    with wt_c2:
        st.subheader(f"🍩 Worktype Allocation ({latest_year})")
        pie_wt = df_wt_agg[(df_wt_agg["Year"] == latest_year) & (df_wt_agg["Period"] == "Actual")]
        fig_wt_pie = px.pie(
            pie_wt, names="Worktype", values="Budget",
            color="Worktype", color_discrete_map=wt_colors, hole=0.45,
        )
        fig_wt_pie.update_traces(
            textinfo="percent+label+value", textposition="outside",
            texttemplate="%{label}<br>%{percent}<br>$%{value:,.0f}"
        )
        fig_wt_pie.update_layout(
            margin=dict(l=20, r=20, t=10, b=20), height=450, showlegend=False,
        )
        st.plotly_chart(fig_wt_pie, use_container_width=True)

    # ── Trend lines per worktype (diverging projections) ──
    st.subheader("📈 Worktype Trends — Diverging Growth Projections")
    fig_wt_trend = go.Figure()
    for wt in wt_categories:
        subset = df_wt_agg[df_wt_agg["Worktype"] == wt]
        actual_w = subset[subset["Period"] == "Actual"].sort_values("Year")
        proj_w = subset[subset["Period"] == "Projected"].sort_values("Year")
        last_a = actual_w[actual_w["Year"] == actual_w["Year"].max()]
        proj_bridge = pd.concat([last_a.assign(Period="Projected"), proj_w])
        color = wt_colors[wt]
        fig_wt_trend.add_trace(go.Scatter(
            x=actual_w["Year"], y=actual_w["Budget"],
            mode="lines+markers", name=wt,
            legendgroup=wt,
            line=dict(color=color, width=2.5), marker=dict(size=7),
        ))
        fig_wt_trend.add_trace(go.Scatter(
            x=proj_bridge["Year"], y=proj_bridge["Budget"],
            mode="lines+markers", name=wt,
            legendgroup=wt, showlegend=False,
            line=dict(color=color, width=2.5, dash="dash"),
            marker=dict(size=7, symbol="diamond"),
        ))
    fig_wt_trend.update_layout(
        yaxis_tickprefix="$", yaxis_tickformat=",",
        xaxis=dict(dtick=1),
        margin=dict(l=20, r=20, t=10, b=20), height=420,
        legend=dict(orientation="h", yanchor="bottom", y=-0.35, font=dict(size=10)),
    )
    st.plotly_chart(fig_wt_trend, use_container_width=True)

    # ── Projected share shift ──
    # ── Overlap Preview ──
    st.subheader("🔀 Worktype Overlap Across Sectors")
    st.caption("Preview of where duplicated worktypes create consolidation opportunities. Full analysis in Savings tab.")
    if len(df_wt_overlap) > 0:
        ov_preview = df_wt_overlap.copy()
        ov_preview["Savings Potential"] = ov_preview["Estimated Overlap (%)"] * ov_preview["Estimated % of Total Org Load (Per Group)"] / 100
        fig_ov_preview = px.scatter(
            ov_preview, x="Estimated Overlap (%)", y="Estimated % of Total Org Load (Per Group)",
            size="Savings Potential", color="Work Type",
            hover_name="Overlap Group",
            color_discrete_map=wt_colors,
            size_max=40,
        )
        fig_ov_preview.update_layout(
            xaxis_title="Overlap %", yaxis_title="Load % (Per Group)",
            margin=dict(l=20, r=20, t=10, b=20), height=400,
            legend=dict(orientation="h", yanchor="bottom", y=-0.25, font=dict(size=10)),
        )
        st.plotly_chart(fig_ov_preview, use_container_width=True)

    # ── Summary table ──
    st.subheader("📋 Worktype Summary Table")
    wt_summary = df_wt_agg.pivot_table(
        index="Worktype", columns="Year", values="Budget", aggfunc="sum"
    )
    wt_summary.columns = [str(int(c)) for c in wt_summary.columns]
    wt_summary["Total"] = wt_summary.sum(axis=1)
    st.dataframe(wt_summary.style.format("${:,.0f}"), use_container_width=True, height=250)

    # ── Department breakdown by worktype ──
    if selected_department != "All Departments":
        st.subheader(f"🏢 {selected_department} — Worktype Breakdown ({latest_year})")
        dept_wt_data = df_wt[
            (df_wt["Department"] == selected_department) &
            (df_wt["Year"] == latest_year) &
            (df_wt["Period"] == "Actual")
        ].groupby("Worktype")["Budget"].sum().reset_index()
        if not dept_wt_data.empty:
            fig_wt_dept = px.bar(
                dept_wt_data, x="Worktype", y="Budget",
                color="Worktype", color_discrete_map=wt_colors,
            )
            fig_wt_dept.update_layout(
                yaxis_tickprefix="$", yaxis_tickformat=",",
                margin=dict(l=20, r=20, t=10, b=20), height=350,
                showlegend=False,
            )
            st.plotly_chart(fig_wt_dept, use_container_width=True)

# ══════════════════════════════════════════════════════════════
# TAB 5: INFRASTRUCTURE BREAKDOWN
# ══════════════════════════════════════════════════════════════
with tab_infra:

    st.caption(
        f"Infrastructure budget ({pct_infra}% of Outside Services) split across 4 categories per department. "
        f"Splits loaded from **worktype_splits.csv**. Growth multiplier: **{growth_multiplier}x**."
    )

    # Get Outside Services by department and year
    os_mask_inf = (
        (df["Department"].isin(selected_departments)) &
        (df["IT Type"].isin(["IT Outside Services", "IT Outside Services*"]))
    )
    os_data_inf = df[os_mask_inf]

    # Build infra category data using department-level splits
    inf_rows = []
    for dept in selected_departments:
        dept_os = os_data_inf[os_data_inf["Department"] == dept].groupby("Year")["Budget"].sum().reset_index()
        dept_os["Budget"] = dept_os["Budget"] * outside_split["Infrastructure"]
        splits = dept_infra_splits.get(dept, {cat: 0.25 for cat in inf_categories})

        for _, row in dept_os.iterrows():
            for cat in inf_categories:
                inf_rows.append({
                    "Year": row["Year"], "Category": cat,
                    "Budget": row["Budget"] * splits[cat], "Period": "Actual",
                    "Department": dept,
                })

        for cat in inf_categories:
            cat_hist = [
                dept_os[dept_os["Year"] == y]["Budget"].sum() * splits[cat]
                for y in all_hist_years
            ]
            compound_multiplier = growth_multiplier * infra_growth[cat]
            cat_proj = project_series(all_hist_years, cat_hist, future_years, compound_multiplier, projection_noise, seed=hash((dept, cat)) % 10000)
            for fy, fv in zip(future_years, cat_proj):
                inf_rows.append({
                    "Year": fy, "Category": cat,
                    "Budget": fv, "Period": "Projected",
                    "Department": dept,
                })

    df_inf = pd.DataFrame(inf_rows)
    df_inf_agg = df_inf.groupby(["Year", "Category", "Period"])["Budget"].sum().reset_index()

    # ── KPIs ──
    inf_latest_by_cat = df_inf_agg[(df_inf_agg["Year"] == latest_year) & (df_inf_agg["Period"] == "Actual")]
    inf_total_latest = inf_latest_by_cat["Budget"].sum()
    ik1, ik2, ik3, ik4 = st.columns(4)
    for col, cat in zip([ik1, ik2, ik3, ik4], inf_categories):
        val = inf_latest_by_cat[inf_latest_by_cat["Category"] == cat]["Budget"].sum()
        pct = (val / inf_total_latest * 100) if inf_total_latest else 0
        col.metric(cat, fmt(val), f"{pct:.1f}% of infra")

    st.divider()

    inf_colors = {
        "Integration": "#1B3A5F",
        "Access": "#1A7A72",
        "Data": "#D4A84B",
        "Communications": "#8B2942",
    }

    inf_c1, inf_c2 = st.columns(2)

    with inf_c1:
        st.subheader("📊 Infrastructure Categories by Year")
        fig_inf_stack = go.Figure()
        for cat in inf_categories:
            subset = df_inf_agg[df_inf_agg["Category"] == cat].sort_values("Year")
            actual_b = subset[subset["Period"] == "Actual"]
            proj_b = subset[subset["Period"] == "Projected"]
            color = inf_colors[cat]
            fig_inf_stack.add_trace(go.Bar(
                x=actual_b["Year"], y=actual_b["Budget"],
                name=cat, legendgroup=cat,
                marker_color=color,
            ))
            fig_inf_stack.add_trace(go.Bar(
                x=proj_b["Year"], y=proj_b["Budget"],
                name=cat, legendgroup=cat, showlegend=False,
                marker_color=color, marker_pattern_shape="/",
            ))
        fig_inf_stack.update_layout(
            barmode="stack",
            yaxis_tickprefix="$", yaxis_tickformat=",",
            xaxis=dict(dtick=1),
            margin=dict(l=20, r=20, t=10, b=20), height=450,
            legend=dict(orientation="h", yanchor="bottom", y=-0.25),
        )
        st.plotly_chart(fig_inf_stack, use_container_width=True)

    with inf_c2:
        st.subheader(f"🍩 Infrastructure Allocation ({latest_year})")
        pie_inf = df_inf_agg[(df_inf_agg["Year"] == latest_year) & (df_inf_agg["Period"] == "Actual")]
        fig_inf_pie = px.pie(
            pie_inf, names="Category", values="Budget",
            color="Category", color_discrete_map=inf_colors, hole=0.45,
        )
        fig_inf_pie.update_traces(
            textinfo="percent+label+value", textposition="outside",
            texttemplate="%{label}<br>%{percent}<br>$%{value:,.0f}"
        )
        fig_inf_pie.update_layout(
            margin=dict(l=20, r=20, t=10, b=20), height=450, showlegend=False,
        )
        st.plotly_chart(fig_inf_pie, use_container_width=True)

    # ── Trend lines with diverging projections ──
    st.subheader("📈 Infrastructure Trends — Diverging Growth Projections")
    fig_inf_trend = go.Figure()
    for cat in inf_categories:
        subset = df_inf_agg[df_inf_agg["Category"] == cat]
        actual_i = subset[subset["Period"] == "Actual"].sort_values("Year")
        proj_i = subset[subset["Period"] == "Projected"].sort_values("Year")
        last_a = actual_i[actual_i["Year"] == actual_i["Year"].max()]
        proj_bridge = pd.concat([last_a.assign(Period="Projected"), proj_i])
        color = inf_colors[cat]
        fig_inf_trend.add_trace(go.Scatter(
            x=actual_i["Year"], y=actual_i["Budget"],
            mode="lines+markers", name=cat,
            legendgroup=cat,
            line=dict(color=color, width=2.5), marker=dict(size=7),
        ))
        fig_inf_trend.add_trace(go.Scatter(
            x=proj_bridge["Year"], y=proj_bridge["Budget"],
            mode="lines+markers", name=cat,
            legendgroup=cat, showlegend=False,
            line=dict(color=color, width=2.5, dash="dash"),
            marker=dict(size=7, symbol="diamond"),
        ))
    fig_inf_trend.update_layout(
        yaxis_tickprefix="$", yaxis_tickformat=",",
        xaxis=dict(dtick=1),
        margin=dict(l=20, r=20, t=10, b=20), height=420,
        legend=dict(orientation="h", yanchor="bottom", y=-0.3, font=dict(size=10)),
    )
    st.plotly_chart(fig_inf_trend, use_container_width=True)

    # ── Projected share shift ──
    # ── Fragmentation Preview ──
    st.subheader("🏗️ Infrastructure Fragmentation Across Sectors")
    st.caption("Preview of where duplicated infrastructure creates consolidation opportunities. Full analysis in Savings tab.")
    if len(df_inf_frag) > 0:
        frag_preview = df_inf_frag.copy()
        frag_preview["Savings Potential"] = frag_preview["Estimated Overlap (%)"] * frag_preview["Estimated % of Total Org Infrastructure Load (Per Group)"] / 100
        fig_frag_preview = px.scatter(
            frag_preview, x="Estimated Overlap (%)", y="Estimated % of Total Org Infrastructure Load (Per Group)",
            size="Savings Potential", color="Infrastructure Type",
            hover_name="Overlap Group",
            color_discrete_map=inf_colors,
            size_max=40,
        )
        fig_frag_preview.update_layout(
            xaxis_title="Overlap %", yaxis_title="Load % (Per Group)",
            margin=dict(l=20, r=20, t=10, b=20), height=400,
            legend=dict(orientation="h", yanchor="bottom", y=-0.25, font=dict(size=10)),
        )
        st.plotly_chart(fig_frag_preview, use_container_width=True)

    # ── Summary table ──
    st.subheader("📋 Infrastructure Summary Table")
    inf_summary = df_inf_agg.pivot_table(
        index="Category", columns="Year", values="Budget", aggfunc="sum"
    )
    inf_summary.columns = [str(int(c)) for c in inf_summary.columns]
    inf_summary["Total"] = inf_summary.sum(axis=1)
    st.dataframe(inf_summary.style.format("${:,.0f}"), use_container_width=True, height=220)

    # ── Department breakdown ──
    if selected_department != "All Departments":
        st.subheader(f"🏢 {selected_department} — Infrastructure Breakdown ({latest_year})")
        dept_inf_data = df_inf[
            (df_inf["Department"] == selected_department) &
            (df_inf["Year"] == latest_year) &
            (df_inf["Period"] == "Actual")
        ].groupby("Category")["Budget"].sum().reset_index()
        if not dept_inf_data.empty:
            fig_inf_dept = px.bar(
                dept_inf_data, x="Category", y="Budget",
                color="Category", color_discrete_map=inf_colors,
            )
            fig_inf_dept.update_layout(
                yaxis_tickprefix="$", yaxis_tickformat=",",
                margin=dict(l=20, r=20, t=10, b=20), height=350,
                showlegend=False,
            )
            st.plotly_chart(fig_inf_dept, use_container_width=True)


# ══════════════════════════════════════════════════════════════
# SHARED: compute savings data across ALL years + projections
# ══════════════════════════════════════════════════════════════

dept_sector_map = df.drop_duplicates("Department").set_index("Department")["Sector"].to_dict()
selected_sectors = set(dept_sector_map.get(d, "") for d in selected_departments)

# Pre-compute overlap/fragmentation ratios (static across years)
wt_overlap_ratios = []
for _, row in df_wt_overlap.iterrows():
    sectors_involved = [s.strip() for s in str(row["Sectors Involved"]).split(",")]
    matching_sectors = [s for s in sectors_involved if s in selected_sectors]
    if not matching_sectors:
        continue
    sector_coverage = len(matching_sectors) / len(sectors_involved)
    wt_overlap_ratios.append({
        "Work Type": row["Work Type"],
        "Overlap Group": row["Overlap Group"],
        "Sectors Involved": ", ".join(sectors_involved),
        "Matching Sectors": len(matching_sectors),
        "Overlap %": row["Estimated Overlap (%)"],
        "Load %": row["Estimated % of Total Org Load (Per Group)"],
        "overlap_frac": row["Estimated Overlap (%)"] / 100,
        "load_frac": row["Estimated % of Total Org Load (Per Group)"] / 100,
        "sector_coverage": sector_coverage,
    })

inf_frag_ratios = []
for _, row in df_inf_frag.iterrows():
    sectors_involved = [s.strip() for s in str(row["Sectors Involved"]).split(",")]
    matching_sectors = [s for s in sectors_involved if s in selected_sectors]
    if not matching_sectors:
        continue
    sector_coverage = len(matching_sectors) / len(sectors_involved)
    inf_frag_ratios.append({
        "Infra Type": row["Infrastructure Type"],
        "Overlap Group": row["Overlap Group"],
        "Sectors Involved": ", ".join(sectors_involved),
        "Matching Sectors": len(matching_sectors),
        "Overlap %": row["Estimated Overlap (%)"],
        "Load %": row["Estimated % of Total Org Infrastructure Load (Per Group)"],
        "overlap_frac": row["Estimated Overlap (%)"] / 100,
        "load_frac": row["Estimated % of Total Org Infrastructure Load (Per Group)"] / 100,
        "sector_coverage": sector_coverage,
    })

# Get OS budget by year (actual)
os_mask_all = (
    (df["Department"].isin(selected_departments)) &
    (df["IT Type"].isin(["IT Outside Services", "IT Outside Services*"]))
)
os_by_year = df[os_mask_all].groupby("Year")["Budget"].sum().reset_index()

# Get SW budget by year (actual)
sw_mask_all = (
    (df["Department"].isin(selected_departments)) &
    (df["IT Type"].str.strip().isin(["IT Software"]))
)
sw_by_year = df[sw_mask_all].groupby("Year")["Budget"].sum().reset_index()

# Project OS and SW budgets into future years
os_hist_vals = [os_by_year[os_by_year["Year"] == y]["Budget"].sum() for y in all_hist_years]
os_proj_vals = project_series(all_hist_years, os_hist_vals, future_years, growth_multiplier, projection_noise, seed=200)

sw_hist_vals = [sw_by_year[sw_by_year["Year"] == y]["Budget"].sum() for y in all_hist_years]
sw_proj_vals = project_series(all_hist_years, sw_hist_vals, future_years, growth_multiplier, projection_noise, seed=300)

# Build year-by-year savings for all years
all_years_range = list(range(selected_years[0], selected_years[1] + 1)) + future_years
savings_by_year = []

# AI replaceability map
replaceability_map = {"High": 0.70, "Medium": 0.35, "Low": 0.10}

for year in all_years_range:
    if year in all_hist_years:
        os_budget = os_by_year[os_by_year["Year"] == year]["Budget"].sum()
        sw_budget = sw_by_year[sw_by_year["Year"] == year]["Budget"].sum()
        period = "Actual"
    else:
        idx = future_years.index(year)
        os_budget = os_proj_vals[idx]
        sw_budget = sw_proj_vals[idx]
        period = "Projected"

    appdev_budget = os_budget * outside_split["Application Development"]
    infra_budget = os_budget * outside_split["Infrastructure"]

    # Worktype overlap savings
    yr_wt_savings = 0
    for r in wt_overlap_ratios:
        yr_wt_savings += appdev_budget * r["load_frac"] * r["sector_coverage"] * r["overlap_frac"] * wt_savings_factor

    # Infra fragmentation savings
    yr_inf_savings = 0
    for r in inf_frag_ratios:
        yr_inf_savings += infra_budget * r["load_frac"] * r["sector_coverage"] * r["overlap_frac"] * inf_savings_factor

    # AI license savings
    yr_ai_savings = 0
    if sw_budget > 0 and len(df_saas) > 0:
        df_tmp = df_saas.copy()
        df_tmp["Weight"] = 1 / df_tmp["Price Per User"]
        df_tmp["Weight"] = df_tmp["Weight"] / df_tmp["Weight"].sum()
        df_tmp["Alloc"] = sw_budget * df_tmp["Weight"]
        df_tmp["Lic"] = (df_tmp["Alloc"] / df_tmp["Price Per User"]).round(0)
        df_tmp["Cost"] = df_tmp["Lic"] * df_tmp["Price Per User"]
        ratio = sw_budget / df_tmp["Cost"].sum() if df_tmp["Cost"].sum() > 0 else 1
        df_tmp["Lic"] = (df_tmp["Lic"] * ratio).round(0)
        df_tmp["Cost"] = df_tmp["Lic"] * df_tmp["Price Per User"]
        df_tmp["ReplPct"] = df_tmp["AI Replaceability"].map(replaceability_map)
        yr_ai_savings = (df_tmp["Cost"] * df_tmp["ReplPct"] * ai_replacement_factor).sum()

    savings_by_year.append({
        "Year": year, "Period": period,
        "App Overlap": yr_wt_savings,
        "Infra Consolidation": yr_inf_savings,
        "AI License Replacement": yr_ai_savings,
        "Total Savings": yr_wt_savings + yr_inf_savings + yr_ai_savings,
        "OS Budget": os_budget,
        "SW Budget": sw_budget,
    })

df_savings_yr = pd.DataFrame(savings_by_year)

# Latest year detail data (for decomposition charts)
total_os_latest = df_savings_yr[df_savings_yr["Year"] == latest_year]["OS Budget"].sum()
total_appdev_budget = total_os_latest * outside_split["Application Development"]
total_infra_budget = total_os_latest * outside_split["Infrastructure"]
total_sw_budget = df_savings_yr[df_savings_yr["Year"] == latest_year]["SW Budget"].sum()

# Build detail dataframes for latest year
df_wt_savings_rows = []
for r in wt_overlap_ratios:
    base = total_appdev_budget * r["load_frac"] * r["sector_coverage"]
    pot = base * r["overlap_frac"] * wt_savings_factor
    df_wt_savings_rows.append({**r, "Base Budget": base, "Potential Savings": pot})
df_wt_savings = pd.DataFrame(df_wt_savings_rows) if df_wt_savings_rows else pd.DataFrame()

df_inf_savings_rows = []
for r in inf_frag_ratios:
    base = total_infra_budget * r["load_frac"] * r["sector_coverage"]
    pot = base * r["overlap_frac"] * inf_savings_factor
    df_inf_savings_rows.append({**r, "Base Budget": base, "Potential Savings": pot})
df_inf_savings = pd.DataFrame(df_inf_savings_rows) if df_inf_savings_rows else pd.DataFrame()

total_wt_savings = df_wt_savings["Potential Savings"].sum() if len(df_wt_savings) else 0
total_inf_savings = df_inf_savings["Potential Savings"].sum() if len(df_inf_savings) else 0

# AI license detail for latest year
df_lic = pd.DataFrame()
total_ai_savings = 0
if total_sw_budget > 0 and len(df_saas) > 0:
    df_lic = df_saas.copy()
    df_lic["Weight"] = 1 / df_lic["Price Per User"]
    df_lic["Weight"] = df_lic["Weight"] / df_lic["Weight"].sum()
    df_lic["Allocated Budget"] = total_sw_budget * df_lic["Weight"]
    df_lic["Est. Licenses"] = (df_lic["Allocated Budget"] / df_lic["Price Per User"]).round(0).astype(int)
    df_lic["Total Cost"] = df_lic["Est. Licenses"] * df_lic["Price Per User"]
    cost_ratio = total_sw_budget / df_lic["Total Cost"].sum() if df_lic["Total Cost"].sum() > 0 else 1
    df_lic["Est. Licenses"] = (df_lic["Est. Licenses"] * cost_ratio).round(0).astype(int)
    df_lic["Total Cost"] = df_lic["Est. Licenses"] * df_lic["Price Per User"]
    df_lic["Replace %"] = df_lic["AI Replaceability"].map(replaceability_map)
    df_lic["AI Savings"] = df_lic["Total Cost"] * df_lic["Replace %"] * ai_replacement_factor
    df_lic["Remaining Cost"] = df_lic["Total Cost"] - df_lic["AI Savings"]
    total_ai_savings = df_lic["AI Savings"].sum()

grand_total_savings = total_wt_savings + total_inf_savings + total_ai_savings

# ══════════════════════════════════════════════════════════════
# TAB 6: SAVINGS (with sub-tabs)
# ══════════════════════════════════════════════════════════════
with tab_savings:

    sub_summary, sub_overlap, sub_infra_consol, sub_ai, sub_lifecycle = st.tabs([
        "📊 Total Summary",
        "🔀 Overlapping Applications",
        "🏗️ Infrastructure Consolidation",
        "🤖 AI vs SaaS Addoption",
        "♻️ Lifecycle Optimization",
    ])

    # Shared colors
    source_colors = {
        "App Overlap": "#1B3A5F",
        "Infra Consolidation": "#1A7A72",
        "AI License Replacement": "#D4A84B",
    }
    wt_colors_tree = {
        "Workflow": "#1B3A5F", "Analysis": "#1A7A72",
        "Collaboration": "#D4A84B", "Transactions": "#8B2942",
        "External Facing": "#7C3AED",
    }
    inf_colors_tree = {
        "Integration": "#1B3A5F", "Access": "#1A7A72",
        "Data": "#D4A84B", "Communications": "#8B2942",
    }
    ai_cat_colors = {
        "Collaboration": "#1B3A5F", "Communications": "#1A7A72",
        "Workflow": "#D4A84B", "Finance": "#8B2942",
        "Analysis": "#7C3AED", "External Facing": "#3B82F6",
        "Access": "#059669", "Security": "#D97706",
        "Infrastructure": "#475569", "Transactions": "#8B2942",
        "Development": "#1A7A72", "Integration": "#1B3A5F",
        "Data": "#D4A84B",
    }
    repl_colors = {"High": "#D97706", "Medium": "#3B82F6", "Low": "#475569"}

    # ════════════════════════════════════════════════════════
    # SUB-TAB 1: TOTAL SUMMARY
    # ════════════════════════════════════════════════════════
    with sub_summary:

        ts1, ts2, ts3, ts4 = st.columns(4)
        total_it_budget_latest = filtered[filtered["Year"] == latest_year]["Budget"].sum()
        savings_pct = (grand_total_savings / total_it_budget_latest * 100) if total_it_budget_latest > 0 else 0
        ts1.metric("Total Savings", fmt(grand_total_savings), f"{savings_pct:.1f}% of IT Budget")
        ts2.metric("App Overlap", fmt(total_wt_savings), f"{wt_savings_factor*100:.0f}% capture")
        ts3.metric("Infra Consolidation", fmt(total_inf_savings), f"{inf_savings_factor*100:.0f}% capture")
        ts4.metric("AI vs SaaS savings", fmt(total_ai_savings), f"{ai_replacement_factor*100:.0f}% capture")

        st.divider()

        # ── Savings trend over time ──
        st.subheader("📈 Savings Trend by Year")
        fig_sav_trend = go.Figure()
        for source, col in [("App Overlap", "App Overlap"), ("Infra Consolidation", "Infra Consolidation"), ("AI License Replacement", "AI License Replacement")]:
            actual = df_savings_yr[df_savings_yr["Period"] == "Actual"]
            proj = df_savings_yr[df_savings_yr["Period"] == "Projected"]
            last_a = actual[actual["Year"] == actual["Year"].max()]
            proj_bridge = pd.concat([last_a.assign(Period="Projected"), proj])
            color = source_colors[source]
            fig_sav_trend.add_trace(go.Scatter(
                x=actual["Year"], y=actual[col],
                mode="lines+markers", name=source,
                legendgroup=source,
                line=dict(color=color, width=2.5), marker=dict(size=7),
            ))
            fig_sav_trend.add_trace(go.Scatter(
                x=proj_bridge["Year"], y=proj_bridge[col],
                mode="lines+markers", name=source,
                legendgroup=source, showlegend=False,
                line=dict(color=color, width=2.5, dash="dash"),
                marker=dict(size=7, symbol="diamond"),
            ))
        fig_sav_trend.update_layout(
            yaxis_tickprefix="$", yaxis_tickformat=",",
            xaxis=dict(dtick=1),
            margin=dict(l=20, r=20, t=10, b=20), height=400,
            legend=dict(orientation="h", yanchor="bottom", y=-0.25, font=dict(size=10)),
        )
        st.plotly_chart(fig_sav_trend, use_container_width=True)

        # ── Stacked bar + cumulative ──
        sav_c1, sav_c2 = st.columns(2)

        with sav_c1:
            st.subheader("📊 Savings Breakdown by Year")
            fig_sav_stack = go.Figure()
            for source, col in [("App Overlap", "App Overlap"), ("Infra Consolidation", "Infra Consolidation"), ("AI License Replacement", "AI License Replacement")]:
                actual = df_savings_yr[df_savings_yr["Period"] == "Actual"]
                proj = df_savings_yr[df_savings_yr["Period"] == "Projected"]
                color = source_colors[source]
                fig_sav_stack.add_trace(go.Bar(
                    x=actual["Year"], y=actual[col],
                    name=source, legendgroup=source, marker_color=color,
                ))
                fig_sav_stack.add_trace(go.Bar(
                    x=proj["Year"], y=proj[col],
                    name=source, legendgroup=source, showlegend=False,
                    marker_color=color, marker_pattern_shape="/",
                ))
            fig_sav_stack.update_layout(
                barmode="stack",
                yaxis_tickprefix="$", yaxis_tickformat=",",
                xaxis=dict(dtick=1),
                margin=dict(l=20, r=20, t=10, b=20), height=400,
                legend=dict(orientation="h", yanchor="bottom", y=-0.25, font=dict(size=10)),
            )
            st.plotly_chart(fig_sav_stack, use_container_width=True)

        with sav_c2:
            st.subheader("📈 Cumulative Savings")
            df_cum = df_savings_yr.copy()
            df_cum["Cumulative"] = df_cum["Total Savings"].cumsum()
            fig_cum = go.Figure()
            actual_cum = df_cum[df_cum["Period"] == "Actual"]
            proj_cum = df_cum[df_cum["Period"] == "Projected"]
            last_a_cum = actual_cum[actual_cum["Year"] == actual_cum["Year"].max()]
            proj_bridge_cum = pd.concat([last_a_cum.assign(Period="Projected"), proj_cum])
            fig_cum.add_trace(go.Scatter(
                x=actual_cum["Year"], y=actual_cum["Cumulative"],
                mode="lines+markers", name="Cumulative",
                fill="tozeroy", line=dict(color="#1B3A5F", width=3),
                marker=dict(size=8),
            ))
            fig_cum.add_trace(go.Scatter(
                x=proj_bridge_cum["Year"], y=proj_bridge_cum["Cumulative"],
                mode="lines+markers", name="Cumulative",
                showlegend=False,
                fill="tozeroy", line=dict(color="#1B3A5F", width=3, dash="dash"),
                marker=dict(size=8, symbol="diamond"),
                fillcolor="rgba(45,90,135,0.15)",
            ))
            fig_cum.update_layout(
                yaxis_tickprefix="$", yaxis_tickformat=",",
                xaxis=dict(dtick=1),
                margin=dict(l=20, r=20, t=10, b=20), height=400,
                showlegend=False,
            )
            st.plotly_chart(fig_cum, use_container_width=True)

        # ── Icicle decomposition for latest year ──
        st.subheader(f"🌳 Savings Decomposition ({latest_year})")
        combined_rows = []
        if len(df_wt_savings) > 0:
            for _, row in df_wt_savings.iterrows():
                combined_rows.append({
                    "Source": "App Overlap", "Category": row["Work Type"],
                    "Item": row["Overlap Group"], "Savings": row["Potential Savings"],
                })
        if len(df_inf_savings) > 0:
            for _, row in df_inf_savings.iterrows():
                combined_rows.append({
                    "Source": "Infra Consolidation", "Category": row["Infra Type"],
                    "Item": row["Overlap Group"], "Savings": row["Potential Savings"],
                })
        if len(df_lic) > 0 and total_ai_savings > 0:
            ai_by_cat = df_lic[df_lic["AI Savings"] > 0].groupby(["Category", "AI Replaceability"])["AI Savings"].sum().reset_index()
            for _, row in ai_by_cat.iterrows():
                combined_rows.append({
                    "Source": "AI License Replacement", "Category": row["Category"],
                    "Item": row["AI Replaceability"], "Savings": row["AI Savings"],
                })
        if combined_rows:
            df_combined = pd.DataFrame(combined_rows)
            df_combined["Total"] = "Total Savings"
            fig_total_icicle = px.icicle(
                df_combined, path=["Total", "Source", "Category", "Item"],
                values="Savings", color="Source", color_discrete_map=source_colors,
            )
            fig_total_icicle.update_traces(
                texttemplate="<b>%{label}</b><br>$%{value:,.0f}",
                hovertemplate="<b>%{label}</b><br>$%{value:,.0f}<extra></extra>",
            )
            fig_total_icicle.update_layout(margin=dict(l=10, r=10, t=10, b=10), height=450)
            st.plotly_chart(fig_total_icicle, use_container_width=True)

        # ── Sankey: Savings flow through taxonomy ──
        st.subheader("🔀 Savings Flow — Sector → Asset Type → Categories → Savings/Remaining")
        fig_sankey = build_savings_sankey(
            df, selected_departments, dept_sector_map, short_name_map,
            outside_split, dept_worktype_splits, dept_infra_splits, wt_categories, inf_categories,
            total_wt_savings, total_inf_savings, total_ai_savings, latest_year
        )
        if fig_sankey:
            st.plotly_chart(fig_sankey, use_container_width=True)
        else:
            st.info("No savings data available for Sankey diagram.")

        # ── Summary table ──
        st.subheader("📋 Savings by Year")
        display_sav = df_savings_yr[["Year", "Period", "App Overlap", "Infra Consolidation", "AI License Replacement", "Total Savings"]].copy()
        st.dataframe(
            display_sav.style.format({
                "App Overlap": "${:,.0f}", "Infra Consolidation": "${:,.0f}",
                "AI License Replacement": "${:,.0f}", "Total Savings": "${:,.0f}",
            }),
            use_container_width=True, height=350
        )

    # ════════════════════════════════════════════════════════
    # SUB-TAB 2: OVERLAPPING APPLICATIONS
    # ════════════════════════════════════════════════════════
    with sub_overlap:
        st.caption(f"Capture factor: **{wt_savings_factor*100:.0f}%**")

        if len(df_wt_savings) > 0:
            wt_sav_by_type = df_wt_savings.groupby("Work Type")["Potential Savings"].sum()
            top_wt = wt_sav_by_type.idxmax()
            top_group = df_wt_savings.sort_values("Potential Savings", ascending=False).iloc[0]["Overlap Group"]
            pct_of_appdev = (total_wt_savings / total_appdev_budget * 100) if total_appdev_budget else 0

            ok1, ok2, ok3, ok4 = st.columns(4)
            ok1.metric("Total Overlap Savings", fmt(total_wt_savings))
            ok2.metric("% of App Dev Budget", f"{pct_of_appdev:.1f}%")
            ok3.metric("Highest Work Type", top_wt)
            ok4.metric("Top Overlap Group", top_group[:35])

            st.divider()

            # Trend
            st.subheader("📈 Overlap Savings Trend")
            fig_ov_trend = go.Figure()
            actual_ov = df_savings_yr[df_savings_yr["Period"] == "Actual"]
            proj_ov = df_savings_yr[df_savings_yr["Period"] == "Projected"]
            last_a_ov = actual_ov[actual_ov["Year"] == actual_ov["Year"].max()]
            proj_bridge_ov = pd.concat([last_a_ov.assign(Period="Projected"), proj_ov])
            fig_ov_trend.add_trace(go.Scatter(
                x=actual_ov["Year"], y=actual_ov["App Overlap"],
                mode="lines+markers", name="App Overlap",
                line=dict(color="#1B3A5F", width=3), marker=dict(size=8),
            ))
            fig_ov_trend.add_trace(go.Scatter(
                x=proj_bridge_ov["Year"], y=proj_bridge_ov["App Overlap"],
                mode="lines+markers", name="App Overlap",
                showlegend=False,
                line=dict(color="#1B3A5F", width=3, dash="dash"),
                marker=dict(size=8, symbol="diamond"),
            ))
            fig_ov_trend.update_layout(
                yaxis_tickprefix="$", yaxis_tickformat=",",
                xaxis=dict(dtick=1),
                margin=dict(l=20, r=20, t=10, b=20), height=350,
                showlegend=False,
            )
            st.plotly_chart(fig_ov_trend, use_container_width=True)

            # Icicle
            tree_rows = []
            for _, row in df_wt_savings.iterrows():
                sectors = [s.strip() for s in row["Sectors Involved"].split(",")]
                sps = row["Potential Savings"] / len(sectors) if sectors else 0
                for sector in sectors:
                    tree_rows.append({"Work Type": row["Work Type"], "Overlap Group": row["Overlap Group"], "Sector": sector, "Savings": sps})
            df_tree = pd.DataFrame(tree_rows)
            df_tree["Total"] = "Application Overlap Savings"

            st.subheader(f"🌳 Decomposition ({latest_year})")
            fig_wt_icicle = px.icicle(df_tree, path=["Total", "Work Type", "Overlap Group", "Sector"], values="Savings", color="Work Type", color_discrete_map=wt_colors_tree)
            fig_wt_icicle.update_traces(texttemplate="<b>%{label}</b><br>$%{value:,.0f}", hovertemplate="<b>%{label}</b><br>$%{value:,.0f}<extra></extra>")
            fig_wt_icicle.update_layout(margin=dict(l=10, r=10, t=10, b=10), height=500)
            st.plotly_chart(fig_wt_icicle, use_container_width=True)

            ov_c1, ov_c2 = st.columns(2)
            with ov_c1:
                st.subheader("📊 Savings by Work Type")
                wt_sav_df = wt_sav_by_type.reset_index(); wt_sav_df.columns = ["Work Type", "Savings"]
                fig_wt_bar = px.bar(wt_sav_df.sort_values("Savings", ascending=True), x="Savings", y="Work Type", orientation="h", color="Work Type", color_discrete_map=wt_colors_tree)
                fig_wt_bar.update_layout(xaxis_tickprefix="$", xaxis_tickformat=",", margin=dict(l=20, r=20, t=10, b=20), height=350, showlegend=False, yaxis_title="")
                st.plotly_chart(fig_wt_bar, use_container_width=True)
            with ov_c2:
                st.subheader("📋 Overlap Detail")
                st.dataframe(
                    df_wt_savings[["Work Type", "Overlap Group", "Sectors Involved", "Overlap %", "Load %", "Base Budget", "Potential Savings"]]
                    .sort_values("Potential Savings", ascending=False)
                    .style.format({"Overlap %": "{:.0f}%", "Load %": "{:.1f}%", "Base Budget": "${:,.0f}", "Potential Savings": "${:,.0f}"}),
                    use_container_width=True, height=350
                )

            # ── Network Graph ──
            st.divider()
            st.subheader("🕸️ Overlap Network")
            wt_net_html = build_pyvis_network(
                df_wt_savings, type_col="Work Type", type_colors=wt_colors_tree,
                title="Overlapping Applications Network"
            )
            components.html(wt_net_html, height=580, scrolling=False)

        else:
            st.info("No worktype overlap groups match the selected sectors.")

    # ════════════════════════════════════════════════════════
    # SUB-TAB 3: INFRASTRUCTURE CONSOLIDATION
    # ════════════════════════════════════════════════════════
    with sub_infra_consol:
        st.caption(f"Capture factor: **{inf_savings_factor*100:.0f}%**")

        if len(df_inf_savings) > 0:
            inf_sav_by_type = df_inf_savings.groupby("Infra Type")["Potential Savings"].sum()
            top_inf = inf_sav_by_type.idxmax()
            top_inf_group = df_inf_savings.sort_values("Potential Savings", ascending=False).iloc[0]["Overlap Group"]
            pct_of_infra = (total_inf_savings / total_infra_budget * 100) if total_infra_budget else 0

            ik1, ik2, ik3, ik4 = st.columns(4)
            ik1.metric("Total Fragmentation Savings", fmt(total_inf_savings))
            ik2.metric("% of Infra Budget", f"{pct_of_infra:.1f}%")
            ik3.metric("Highest Infra Type", top_inf)
            ik4.metric("Top Overlap Group", top_inf_group[:35])

            st.divider()

            # Trend
            st.subheader("📈 Consolidation Savings Trend")
            fig_ic_trend = go.Figure()
            actual_ic = df_savings_yr[df_savings_yr["Period"] == "Actual"]
            proj_ic = df_savings_yr[df_savings_yr["Period"] == "Projected"]
            last_a_ic = actual_ic[actual_ic["Year"] == actual_ic["Year"].max()]
            proj_bridge_ic = pd.concat([last_a_ic.assign(Period="Projected"), proj_ic])
            fig_ic_trend.add_trace(go.Scatter(
                x=actual_ic["Year"], y=actual_ic["Infra Consolidation"],
                mode="lines+markers", name="Infra Consolidation",
                line=dict(color="#0F2744", width=3), marker=dict(size=8),
            ))
            fig_ic_trend.add_trace(go.Scatter(
                x=proj_bridge_ic["Year"], y=proj_bridge_ic["Infra Consolidation"],
                mode="lines+markers", name="Infra Consolidation",
                showlegend=False,
                line=dict(color="#0F2744", width=3, dash="dash"),
                marker=dict(size=8, symbol="diamond"),
            ))
            fig_ic_trend.update_layout(
                yaxis_tickprefix="$", yaxis_tickformat=",",
                xaxis=dict(dtick=1),
                margin=dict(l=20, r=20, t=10, b=20), height=350,
                showlegend=False,
            )
            st.plotly_chart(fig_ic_trend, use_container_width=True)

            # Icicle
            inf_tree_rows = []
            for _, row in df_inf_savings.iterrows():
                sectors = [s.strip() for s in row["Sectors Involved"].split(",")]
                sps = row["Potential Savings"] / len(sectors) if sectors else 0
                for sector in sectors:
                    inf_tree_rows.append({"Infra Type": row["Infra Type"], "Overlap Group": row["Overlap Group"], "Sector": sector, "Savings": sps})
            df_inf_tree = pd.DataFrame(inf_tree_rows)
            df_inf_tree["Total"] = "Infrastructure Consolidation Savings"

            st.subheader(f"🌳 Decomposition ({latest_year})")
            fig_inf_icicle = px.icicle(df_inf_tree, path=["Total", "Infra Type", "Overlap Group", "Sector"], values="Savings", color="Infra Type", color_discrete_map=inf_colors_tree)
            fig_inf_icicle.update_traces(texttemplate="<b>%{label}</b><br>$%{value:,.0f}", hovertemplate="<b>%{label}</b><br>$%{value:,.0f}<extra></extra>")
            fig_inf_icicle.update_layout(margin=dict(l=10, r=10, t=10, b=10), height=500)
            st.plotly_chart(fig_inf_icicle, use_container_width=True)

            ic_c1, ic_c2 = st.columns(2)
            with ic_c1:
                st.subheader("📊 Savings by Infra Type")
                inf_sav_df = inf_sav_by_type.reset_index(); inf_sav_df.columns = ["Infra Type", "Savings"]
                fig_inf_bar = px.bar(inf_sav_df.sort_values("Savings", ascending=True), x="Savings", y="Infra Type", orientation="h", color="Infra Type", color_discrete_map=inf_colors_tree)
                fig_inf_bar.update_layout(xaxis_tickprefix="$", xaxis_tickformat=",", margin=dict(l=20, r=20, t=10, b=20), height=350, showlegend=False, yaxis_title="")
                st.plotly_chart(fig_inf_bar, use_container_width=True)
            with ic_c2:
                st.subheader("📋 Fragmentation Detail")
                st.dataframe(
                    df_inf_savings[["Infra Type", "Overlap Group", "Sectors Involved", "Overlap %", "Load %", "Base Budget", "Potential Savings"]]
                    .sort_values("Potential Savings", ascending=False)
                    .style.format({"Overlap %": "{:.0f}%", "Load %": "{:.1f}%", "Base Budget": "${:,.0f}", "Potential Savings": "${:,.0f}"}),
                    use_container_width=True, height=350
                )

            # ── Network Graph ──
            st.divider()
            st.subheader("🕸️ Fragmentation Network")
            inf_net_html = build_pyvis_network(
                df_inf_savings, type_col="Infra Type", type_colors=inf_colors_tree,
                title="Infrastructure Consolidation Network"
            )
            components.html(inf_net_html, height=580, scrolling=False)

        else:
            st.info("No infrastructure fragmentation groups match the selected sectors.")

    # ════════════════════════════════════════════════════════
    # SUB-TAB 4: AI LICENSE REPLACEMENT
    # ════════════════════════════════════════════════════════
    with sub_ai:
        st.caption(f"Capture factor: **{ai_replacement_factor*100:.0f}%**")

        if total_sw_budget > 0 and len(df_lic) > 0:
            total_lic_cost = df_lic["Total Cost"].sum()
            pct_savings = (total_ai_savings / total_lic_cost * 100) if total_lic_cost else 0
            total_licenses = df_lic["Est. Licenses"].sum()
            replaceable_licenses = (df_lic["Est. Licenses"] * df_lic["Replace %"] * ai_replacement_factor).sum()

            ak1, ak2, ak3, ak4 = st.columns(4)
            ak1.metric("Software Budget", fmt(total_sw_budget))
            ak2.metric("AI vs SaaS Trend Savings", fmt(total_ai_savings), f"{pct_savings:.1f}% of software")
            ak3.metric("Est. Total Licenses", f"{total_licenses:,.0f}")
            ak4.metric("Replaceable Licenses", f"{replaceable_licenses:,.0f}")

            st.divider()

            # Trend
            st.subheader("📈 AI vs SaaS Savings Trend")
            fig_ai_trend = go.Figure()
            actual_ai = df_savings_yr[df_savings_yr["Period"] == "Actual"]
            proj_ai = df_savings_yr[df_savings_yr["Period"] == "Projected"]
            last_a_ai = actual_ai[actual_ai["Year"] == actual_ai["Year"].max()]
            proj_bridge_ai = pd.concat([last_a_ai.assign(Period="Projected"), proj_ai])
            fig_ai_trend.add_trace(go.Scatter(
                x=actual_ai["Year"], y=actual_ai["AI License Replacement"],
                mode="lines+markers", name="AI License",
                line=dict(color="#D4A84B", width=3), marker=dict(size=8),
            ))
            fig_ai_trend.add_trace(go.Scatter(
                x=proj_bridge_ai["Year"], y=proj_bridge_ai["AI License Replacement"],
                mode="lines+markers", name="AI License",
                showlegend=False,
                line=dict(color="#D4A84B", width=3, dash="dash"),
                marker=dict(size=8, symbol="diamond"),
            ))
            fig_ai_trend.update_layout(
                yaxis_tickprefix="$", yaxis_tickformat=",",
                xaxis=dict(dtick=1),
                margin=dict(l=20, r=20, t=10, b=20), height=350,
                showlegend=False,
            )
            st.plotly_chart(fig_ai_trend, use_container_width=True)

            # Icicle
            st.subheader(f"🌳 AI Addoption Decomposition ({latest_year})")
            tree_data = df_lic[df_lic["AI Savings"] > 0].copy()
            tree_data["Total"] = "AI License Savings"
            fig_ai_icicle = px.icicle(tree_data, path=["Total", "Category", "AI Replaceability", "License"], values="AI Savings", color="Category", color_discrete_map=ai_cat_colors)
            fig_ai_icicle.update_traces(texttemplate="<b>%{label}</b><br>$%{value:,.0f}", hovertemplate="<b>%{label}</b><br>$%{value:,.0f}<extra></extra>")
            fig_ai_icicle.update_layout(margin=dict(l=10, r=10, t=10, b=10), height=500)
            st.plotly_chart(fig_ai_icicle, use_container_width=True)

            st.divider()

            ai_c1, ai_c2 = st.columns(2)
            with ai_c1:
                st.subheader("📊 Savings by Category")
                cat_savings = df_lic.groupby("Category")[["Total Cost", "AI Savings"]].sum().reset_index()
                cat_savings = cat_savings.sort_values("AI Savings", ascending=True)
                fig_cat = go.Figure()
                fig_cat.add_trace(go.Bar(y=cat_savings["Category"], x=cat_savings["Total Cost"] - cat_savings["AI Savings"], name="Remaining Cost", orientation="h", marker_color="#1A7A72"))
                fig_cat.add_trace(go.Bar(y=cat_savings["Category"], x=cat_savings["AI Savings"], name="AI Savings", orientation="h", marker_color="#D97706"))
                fig_cat.update_layout(barmode="stack", xaxis_tickprefix="$", xaxis_tickformat=",", margin=dict(l=20, r=20, t=10, b=20), height=400, legend=dict(orientation="h", yanchor="bottom", y=-0.2), yaxis_title="")
                st.plotly_chart(fig_cat, use_container_width=True)
            with ai_c2:
                st.subheader("🍩 Savings by Replaceability")
                repl_savings = df_lic.groupby("AI Replaceability")[["AI Savings", "Total Cost"]].sum().reset_index()
                fig_repl = px.pie(repl_savings, names="AI Replaceability", values="AI Savings", color="AI Replaceability", color_discrete_map=repl_colors, hole=0.45)
                fig_repl.update_traces(textinfo="percent+label+value", textposition="outside", texttemplate="%{label}<br>%{percent}<br>$%{value:,.0f}")
                fig_repl.update_layout(margin=dict(l=20, r=20, t=10, b=20), height=400, showlegend=False)
                st.plotly_chart(fig_repl, use_container_width=True)

            st.subheader("🏆 Top 15 Licenses by AI Savings Potential")
            top_lic = df_lic.nlargest(15, "AI Savings")
            fig_top = px.bar(top_lic.sort_values("AI Savings", ascending=True), x="AI Savings", y="License", orientation="h", color="AI Replaceability", color_discrete_map=repl_colors)
            fig_top.update_layout(xaxis_tickprefix="$", xaxis_tickformat=",", margin=dict(l=20, r=20, t=10, b=20), height=450, legend=dict(orientation="h", yanchor="bottom", y=-0.15), yaxis_title="")
            st.plotly_chart(fig_top, use_container_width=True)

            st.subheader("📋 License Detail")
            display_lic = df_lic[["License", "Category", "Price Per User", "AI Replaceability", "Replaceability Reason", "Est. Licenses", "Total Cost", "AI Savings", "Remaining Cost"]].sort_values("AI Savings", ascending=False)
            st.dataframe(
                display_lic.style.format({"Price Per User": "${:,.0f}", "Est. Licenses": "{:,.0f}", "Total Cost": "${:,.0f}", "AI Savings": "${:,.0f}", "Remaining Cost": "${:,.0f}"}),
                use_container_width=True, height=450
            )
        else:
            st.info("No IT Software budget found for the selected departments/year.")

    # ════════════════════════════════════════════════════════
    # SUB-TAB 5: LIFECYCLE OPTIMIZATION
    # ════════════════════════════════════════════════════════
    with sub_lifecycle:
        st.header("♻️ Lifecycle Optimization")
        st.info("🚧 **Coming Soon** — This section will analyze asset lifecycle stages, replacement timelines, and modernization opportunities across the digital portfolio.")

# ══════════════════════════════════════════════════════════════
# TAB 7: LIFECYCLE OPTIMIZATION
# ══════════════════════════════════════════════════════════════
with tab_lifecycle:
    st.header("♻️ Lifecycle Optimization")
    st.info("🚧 **Coming Soon** — This section will provide comprehensive lifecycle analysis including asset age tracking, technology debt assessment, replacement planning, and ROI-based modernization roadmaps.")

st.caption("Data source: City of Chicago IT Budget | Digital Asset Model built with Streamlit & Plotly")
